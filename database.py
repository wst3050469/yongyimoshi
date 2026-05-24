"""
永颐无机磨石 - SQLite数据库模块
提供项目管理和数据持久化支持
"""

import sqlite3
import json
import os
from datetime import datetime
from typing import Dict, List, Optional, Any

DB_DIR = os.path.join(os.path.dirname(__file__), 'data')
DB_PATH = os.path.join(DB_DIR, 'yongyi.db')


def get_db() -> sqlite3.Connection:
    """获取数据库连接"""
    os.makedirs(DB_DIR, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def init_db():
    """初始化数据库表结构"""
    conn = get_db()
    cursor = conn.cursor()

    cursor.executescript("""
        CREATE TABLE IF NOT EXISTS projects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL DEFAULT '默认项目',
            area REAL NOT NULL DEFAULT 100,
            base_thickness REAL NOT NULL DEFAULT 50,
            surface_thickness REAL NOT NULL DEFAULT 15,
            start_date TEXT NOT NULL DEFAULT (date('now')),
            status TEXT NOT NULL DEFAULT '进行中',
            location TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            created_at TEXT NOT NULL DEFAULT (datetime('now')),
            updated_at TEXT NOT NULL DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS daily_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            log_date TEXT NOT NULL,
            weather TEXT DEFAULT '晴',
            temperature TEXT DEFAULT '20~25',
            workers INTEGER DEFAULT 5,
            work_content TEXT DEFAULT '',
            materials_used TEXT DEFAULT '',
            issues TEXT DEFAULT '无',
            created_at TEXT NOT NULL DEFAULT (datetime('now')),
            FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS quality_tests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            test_id TEXT NOT NULL,
            test_name TEXT NOT NULL,
            standard_value TEXT NOT NULL,
            actual_value TEXT DEFAULT '',
            is_pass INTEGER DEFAULT 0,
            test_date TEXT DEFAULT (date('now')),
            FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS checklist_state (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            category TEXT NOT NULL,
            item_id TEXT NOT NULL,
            item_text TEXT NOT NULL,
            is_checked INTEGER DEFAULT 0,
            phase TEXT DEFAULT '',
            FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS photos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            phase TEXT NOT NULL DEFAULT '',
            description TEXT DEFAULT '',
            filename TEXT NOT NULL,
            file_size INTEGER DEFAULT 0,
            created_at TEXT NOT NULL DEFAULT (datetime('now')),
            FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS material_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            material_name TEXT NOT NULL,
            quantity_kg REAL DEFAULT 0,
            quantity_packages REAL DEFAULT 0,
            unit_price REAL DEFAULT 0,
            total_cost REAL DEFAULT 0,
            record_date TEXT DEFAULT (date('now')),
            FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS suppliers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            contact_person TEXT DEFAULT '',
            phone TEXT DEFAULT '',
            address TEXT DEFAULT '',
            materials TEXT DEFAULT '[]',
            rating INTEGER DEFAULT 3,
            notes TEXT DEFAULT '',
            created_at TEXT NOT NULL DEFAULT (datetime('now')),
            updated_at TEXT NOT NULL DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS supplier_prices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            supplier_id INTEGER NOT NULL,
            material_name TEXT NOT NULL,
            unit_price REAL NOT NULL DEFAULT 0,
            price_date TEXT NOT NULL DEFAULT (date('now')),
            notes TEXT DEFAULT '',
            FOREIGN KEY (supplier_id) REFERENCES suppliers(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS curing_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            record_date TEXT NOT NULL DEFAULT (date('now')),
            weather TEXT DEFAULT '晴',
            temp_min REAL DEFAULT 20,
            temp_max REAL DEFAULT 25,
            humidity REAL DEFAULT 60,
            wind_level TEXT DEFAULT '',
            curing_measure TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            created_at TEXT NOT NULL DEFAULT (datetime('now')),
            FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS teams (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            leader TEXT DEFAULT '',
            member_count INTEGER DEFAULT 0,
            specialty TEXT DEFAULT '',
            created_at TEXT NOT NULL DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS workers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            phone TEXT DEFAULT '',
            role TEXT DEFAULT '工人',
            team_id INTEGER DEFAULT 0,
            hourly_rate REAL DEFAULT 0,
            notes TEXT DEFAULT '',
            created_at TEXT NOT NULL DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS work_hours (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            worker_id INTEGER NOT NULL,
            work_date TEXT NOT NULL,
            hours REAL DEFAULT 8,
            work_type TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            created_at TEXT NOT NULL DEFAULT (datetime('now')),
            FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS budget_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            category TEXT NOT NULL,
            planned_amount REAL DEFAULT 0,
            actual_amount REAL DEFAULT 0,
            description TEXT DEFAULT '',
            created_at TEXT NOT NULL DEFAULT (datetime('now')),
            FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS notifications (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            type TEXT NOT NULL,
            title TEXT NOT NULL,
            message TEXT DEFAULT '',
            is_read INTEGER DEFAULT 0,
            due_date TEXT DEFAULT '',
            created_at TEXT NOT NULL DEFAULT (datetime('now')),
            FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS equipment (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            type TEXT DEFAULT '',
            model TEXT DEFAULT '',
            quantity INTEGER DEFAULT 1,
            unit TEXT DEFAULT '台',
            status TEXT DEFAULT '可用',
            purchase_date TEXT DEFAULT '',
            next_maintenance TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            created_at TEXT NOT NULL DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS equipment_usage (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            equipment_id INTEGER NOT NULL,
            project_id INTEGER NOT NULL,
            quantity_used INTEGER DEFAULT 1,
            start_date TEXT NOT NULL,
            end_date TEXT DEFAULT '',
            operator TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            created_at TEXT NOT NULL DEFAULT (datetime('now')),
            FOREIGN KEY (equipment_id) REFERENCES equipment(id) ON DELETE CASCADE,
            FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS acceptance_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            acceptance_type TEXT NOT NULL,
            check_date TEXT NOT NULL,
            inspector TEXT DEFAULT '',
            result TEXT DEFAULT '待定',
            defects TEXT DEFAULT '',
            score REAL DEFAULT 0,
            attachments TEXT DEFAULT '[]',
            notes TEXT DEFAULT '',
            created_at TEXT NOT NULL DEFAULT (datetime('now')),
            FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS acceptance_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            acceptance_id INTEGER NOT NULL,
            item_name TEXT NOT NULL,
            standard TEXT DEFAULT '',
            is_pass INTEGER DEFAULT 0,
            actual_value TEXT DEFAULT '',
            FOREIGN KEY (acceptance_id) REFERENCES acceptance_records(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS safety_checks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            check_date TEXT NOT NULL,
            inspector TEXT DEFAULT '',
            check_type TEXT DEFAULT '日常检查',
            items TEXT DEFAULT '[]',
            total_items INTEGER DEFAULT 0,
            passed_items INTEGER DEFAULT 0,
            result TEXT DEFAULT '合格',
            rectification TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            created_at TEXT NOT NULL DEFAULT (datetime('now')),
            FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS safety_incidents (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            incident_date TEXT NOT NULL,
            incident_type TEXT DEFAULT '其他',
            severity TEXT DEFAULT '轻微',
            description TEXT DEFAULT '',
            injured_person TEXT DEFAULT '',
            treatment TEXT DEFAULT '',
            root_cause TEXT DEFAULT '',
            preventive_measures TEXT DEFAULT '',
            is_closed INTEGER DEFAULT 0,
            closed_date TEXT DEFAULT '',
            created_at TEXT NOT NULL DEFAULT (datetime('now')),
            FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS documents (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER DEFAULT 0,
            doc_type TEXT DEFAULT 'other',
            title TEXT NOT NULL,
            description TEXT DEFAULT '',
            filename TEXT NOT NULL,
            file_size INTEGER DEFAULT 0,
            version TEXT DEFAULT 'v1.0',
            tags TEXT DEFAULT '[]',
            created_at TEXT NOT NULL DEFAULT (datetime('now')),
            FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS material_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            request_date TEXT NOT NULL,
            applicant TEXT DEFAULT '',
            status TEXT DEFAULT '待审批',
            approver TEXT DEFAULT '',
            approved_date TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            created_at TEXT NOT NULL DEFAULT (datetime('now')),
            FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS request_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            request_id INTEGER NOT NULL,
            material_name TEXT NOT NULL,
            quantity_kg REAL DEFAULT 0,
            unit TEXT DEFAULT 'kg',
            purpose TEXT DEFAULT '',
            FOREIGN KEY (request_id) REFERENCES material_requests(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS subcontractors (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            contact_person TEXT DEFAULT '',
            phone TEXT DEFAULT '',
            scope TEXT DEFAULT '',
            contract_amount REAL DEFAULT 0,
            contract_date TEXT DEFAULT '',
            start_date TEXT DEFAULT '',
            end_date TEXT DEFAULT '',
            status TEXT DEFAULT '进行中',
            rating INTEGER DEFAULT 3,
            notes TEXT DEFAULT '',
            created_at TEXT NOT NULL DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS project_templates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            description TEXT DEFAULT '',
            config TEXT NOT NULL DEFAULT '{}',
            created_at TEXT NOT NULL DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS material_consumption (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            material_name TEXT NOT NULL,
            planned_kg REAL DEFAULT 0,
            actual_kg REAL DEFAULT 0,
            unit TEXT DEFAULT 'kg',
            notes TEXT DEFAULT '',
            FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            display_name TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'worker',
            phone TEXT DEFAULT '',
            email TEXT DEFAULT '',
            is_active INTEGER DEFAULT 1,
            last_login TEXT DEFAULT '',
            created_at TEXT NOT NULL DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS environment_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            record_date TEXT NOT NULL DEFAULT (date('now')),
            record_time TEXT DEFAULT '',
            temperature REAL DEFAULT 0,
            humidity REAL DEFAULT 0,
            base_moisture REAL DEFAULT NULL,
            surface_temp REAL DEFAULT NULL,
            wind_speed REAL DEFAULT NULL,
            weather_condition TEXT DEFAULT '',
            recorder TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            created_at TEXT NOT NULL DEFAULT (datetime('now')),
            FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS project_phases (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            phase_name TEXT NOT NULL,
            phase_order INTEGER NOT NULL,
            status TEXT NOT NULL DEFAULT 'pending',
            started_at TEXT DEFAULT '',
            completed_at TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            created_at TEXT NOT NULL DEFAULT (datetime('now')),
            FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE
        );
    """)

    conn.commit()
    conn.close()


# ============================================================
# 项目管理
# ============================================================

def create_project(name: str = "新建项目", area: float = 100,
                   base_thickness: float = 50, surface_thickness: float = 15,
                   start_date: str = "", location: str = "") -> int:
    """创建新项目"""
    if not start_date:
        start_date = datetime.now().strftime("%Y-%m-%d")
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO projects (name, area, base_thickness, surface_thickness, start_date, location)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (name, area, base_thickness, surface_thickness, start_date, location))
    pid = cursor.lastrowid

    # 为新项目初始化检查清单
    _init_checklist_for_project(pid, conn)
    # 初始化质量检测项
    _init_quality_tests_for_project(pid, conn)
    # 初始化工序状态
    _init_project_phases(pid, conn)

    conn.commit()
    conn.close()
    return pid


def _init_checklist_for_project(pid: int, conn: sqlite3.Connection):
    """为新项目初始化检查清单"""
    from materials_calc import CHECKLIST_ITEMS
    cursor = conn.cursor()
    for cat in CHECKLIST_ITEMS:
        for item in cat["items"]:
            cursor.execute("""
                INSERT INTO checklist_state (project_id, category, item_id, item_text, phase)
                VALUES (?, ?, ?, ?, ?)
            """, (pid, cat["category"], item["id"], item["text"], cat.get("phase", "")))


def _init_quality_tests_for_project(pid: int, conn: sqlite3.Connection):
    """为新项目初始化质量检测项"""
    from materials_calc import QUALITY_TESTS
    cursor = conn.cursor()
    for t in QUALITY_TESTS:
        cursor.execute("""
            INSERT INTO quality_tests (project_id, test_id, test_name, standard_value)
            VALUES (?, ?, ?, ?)
        """, (pid, t["id"], t["name"], t["standard"]))


def _init_project_phases(pid: int, conn: sqlite3.Connection):
    """为新项目初始化工序状态"""
    cursor = conn.cursor()
    cursor.execute("DELETE FROM project_phases WHERE project_id = ?", (pid,))
    for i, name in enumerate(PHASE_ORDER, 1):
        cursor.execute("""
            INSERT INTO project_phases (project_id, phase_name, phase_order, status)
            VALUES (?, ?, ?, ?)
        """, (pid, name, i, 'pending'))


def get_projects() -> List[Dict]:
    """获取所有项目"""
    conn = get_db()
    cursor = conn.cursor()
    rows = cursor.execute("""
        SELECT p.*,
            (SELECT COUNT(*) FROM checklist_state WHERE project_id = p.id AND is_checked = 1) as checked_items,
            (SELECT COUNT(*) FROM checklist_state WHERE project_id = p.id) as total_items,
            (SELECT COUNT(*) FROM daily_logs WHERE project_id = p.id) as log_count
        FROM projects p ORDER BY p.updated_at DESC
    """).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_project(pid: int) -> Optional[Dict]:
    """获取单个项目"""
    conn = get_db()
    cursor = conn.cursor()
    row = cursor.execute("SELECT * FROM projects WHERE id = ?", (pid,)).fetchone()
    conn.close()
    return dict(row) if row else None


def update_project(pid: int, **kwargs) -> bool:
    """更新项目信息"""
    allowed = ['name', 'area', 'base_thickness', 'surface_thickness',
               'start_date', 'status', 'location', 'notes']
    updates = {k: v for k, v in kwargs.items() if k in allowed}
    if not updates:
        return False
    updates['updated_at'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    set_clause = ", ".join(f"{k} = ?" for k in updates)
    values = list(updates.values()) + [pid]
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(f"UPDATE projects SET {set_clause} WHERE id = ?", values)
    conn.commit()
    conn.close()
    return True


def delete_project(pid: int) -> bool:
    """删除项目"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM projects WHERE id = ?", (pid,))
    conn.commit()
    conn.close()
    return True


# ============================================================
# 日志管理
# ============================================================

def save_daily_log(project_id: int, data: Dict) -> int:
    """保存施工日志"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO daily_logs (project_id, log_date, weather, temperature,
            workers, work_content, materials_used, issues)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (project_id, data.get('date', ''), data.get('weather', '晴'),
          data.get('temperature', ''), int(data.get('workers', 5)),
          data.get('work_content', ''), data.get('materials_used', ''),
          data.get('issues', '无')))
    log_id = cursor.lastrowid
    conn.commit()
    conn.close()
    return log_id


def get_daily_logs(project_id: int) -> List[Dict]:
    """获取项目的施工日志"""
    conn = get_db()
    cursor = conn.cursor()
    rows = cursor.execute("""
        SELECT * FROM daily_logs WHERE project_id = ? ORDER BY log_date DESC
    """, (project_id,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ============================================================
# 检查清单
# ============================================================

def get_checklist(project_id: int) -> List[Dict]:
    """获取项目的检查清单（按category分组）"""
    conn = get_db()
    cursor = conn.cursor()
    rows = cursor.execute("""
        SELECT * FROM checklist_state WHERE project_id = ? ORDER BY id
    """, (project_id,)).fetchall()
    conn.close()

    # 按category分组
    categories = {}
    for r in rows:
        d = dict(r)
        cat = d['category']
        if cat not in categories:
            categories[cat] = {"category": cat, "phase": d['phase'], "items": []}
        categories[cat]["items"].append({
            "id": d['item_id'],
            "text": d['item_text'],
            "checked": bool(d['is_checked']),
        })
    return list(categories.values())


def update_checklist(project_id: int, checklist_data: List[Dict]):
    """更新检查清单状态"""
    conn = get_db()
    cursor = conn.cursor()
    for cat in checklist_data:
        for item in cat.get('items', []):
            cursor.execute("""
                UPDATE checklist_state SET is_checked = ? 
                WHERE project_id = ? AND item_id = ?
            """, (1 if item.get('checked') else 0, project_id, item.get('id')))
    conn.commit()
    conn.close()


# ============================================================
# 质量检测
# ============================================================

def get_quality_tests(project_id: int) -> Dict:
    """获取项目的质量检测数据"""
    conn = get_db()
    cursor = conn.cursor()
    rows = cursor.execute("""
        SELECT * FROM quality_tests WHERE project_id = ? ORDER BY id
    """, (project_id,)).fetchall()
    conn.close()

    from materials_calc import QUALITY_TESTS
    results = {}
    for r in rows:
        d = dict(r)
        results[d['test_id']] = d['actual_value']
        results[d['test_id'] + '_pass'] = bool(d['is_pass'])

    return {
        "templates": QUALITY_TESTS,
        "results": results,
    }


def save_quality_tests(project_id: int, results: Dict):
    """保存质量检测结果"""
    conn = get_db()
    cursor = conn.cursor()
    from materials_calc import QUALITY_TESTS
    for t in QUALITY_TESTS:
        val = results.get(t['id'], '')
        is_pass = 1 if results.get(t['id'] + '_pass', False) else 0
        cursor.execute("""
            UPDATE quality_tests SET actual_value = ?, is_pass = ?, test_date = date('now')
            WHERE project_id = ? AND test_id = ?
        """, (val, is_pass, project_id, t['id']))
    conn.commit()
    conn.close()


# ============================================================
# 项目看板数据
# ============================================================

def get_dashboard_data(project_id: int) -> Dict:
    """获取项目看板数据"""
    conn = get_db()
    cursor = conn.cursor()

    # 检查清单进度
    rows = cursor.execute("""
        SELECT phase, COUNT(*) as total,
               SUM(CASE WHEN is_checked = 1 THEN 1 ELSE 0 END) as checked
        FROM checklist_state WHERE project_id = ?
        GROUP BY phase
    """, (project_id,)).fetchall()

    total_items = 0
    checked_items = 0
    phase_progress = {}
    for r in rows:
        d = dict(r)
        total = d['total']
        checked = d['checked'] or 0
        total_items += total
        checked_items += checked
        phase_progress[d['phase']] = {
            "total": total,
            "checked": checked,
            "percent": round(checked / total * 100, 1) if total > 0 else 0,
        }

    overall_pct = round(checked_items / total_items * 100, 1) if total_items > 0 else 0

    # 日志数量
    log_count = cursor.execute(
        "SELECT COUNT(*) FROM daily_logs WHERE project_id = ?", (project_id,)
    ).fetchone()[0]

    conn.close()

    return {
        "overall_progress": overall_pct,
        "total_items": total_items,
        "checked_items": checked_items,
        "phase_progress": phase_progress,
        "log_count": log_count,
    }


# ============================================================
# 项目报告
# ============================================================

def generate_report(project_id: int) -> Dict:
    """生成项目综合报告"""
    project = get_project(project_id)
    if not project:
        return {"error": "项目不存在"}

    dashboard = get_dashboard_data(project_id)
    logs = get_daily_logs(project_id)
    quality = get_quality_tests(project_id)
    checklist = get_checklist(project_id)

    # 材料用量估算
    from materials_calc import calculate_all
    calc = calculate_all(project['area'], project['base_thickness'], project['surface_thickness'])

    return {
        "project": project,
        "dashboard": dashboard,
        "logs": logs,
        "quality": quality,
        "materials": calc,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


# 初始化数据库
init_db()


# ============================================================
# 材料库存管理
# ============================================================

def add_material_record(project_id: int, material_name: str,
                        quantity_kg: float, unit_price: float = 0,
                        record_type: str = "入库",
                        record_date: str = "") -> int:
    """添加材料入库/出库记录"""
    if not record_date:
        record_date = datetime.now().strftime("%Y-%m-%d")
    packages = 0
    total_cost = quantity_kg * unit_price
    from materials_calc import PACKAGING_SPECS
    spec = PACKAGING_SPECS.get(material_name)
    if spec:
        import math
        packages = math.ceil(quantity_kg / spec["per_package"])

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO material_records 
        (project_id, material_name, quantity_kg, quantity_packages, 
         unit_price, total_cost, record_date)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (project_id, material_name, quantity_kg, packages,
          unit_price, total_cost, record_date))
    rid = cursor.lastrowid
    conn.commit()
    conn.close()
    return rid


def get_material_inventory(project_id: int) -> List[Dict]:
    """获取项目材料库存汇总"""
    conn = get_db()
    cursor = conn.cursor()
    rows = cursor.execute("""
        SELECT material_name,
               SUM(quantity_kg) as total_kg,
               SUM(quantity_packages) as total_packages,
               SUM(total_cost) as total_cost,
               COUNT(*) as record_count
        FROM material_records
        WHERE project_id = ?
        GROUP BY material_name
        ORDER BY material_name
    """, (project_id,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_material_records(project_id: int, limit: int = 50) -> List[Dict]:
    """获取项目的材料进出记录"""
    conn = get_db()
    cursor = conn.cursor()
    rows = cursor.execute("""
        SELECT * FROM material_records
        WHERE project_id = ?
        ORDER BY record_date DESC, id DESC
        LIMIT ?
    """, (project_id, limit)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ============================================================
# 供应商管理
# ============================================================

def add_supplier(name: str, contact_person: str = "", phone: str = "",
                 address: str = "", materials: str = "[]",
                 rating: int = 3, notes: str = "") -> int:
    """添加供应商"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO suppliers (name, contact_person, phone, address, materials, rating, notes)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (name, contact_person, phone, address, materials, rating, notes))
    sid = cursor.lastrowid
    conn.commit()
    conn.close()
    return sid


def get_suppliers() -> List[Dict]:
    """获取所有供应商"""
    conn = get_db()
    cursor = conn.cursor()
    rows = cursor.execute("""
        SELECT s.*,
            (SELECT COUNT(*) FROM supplier_prices WHERE supplier_id = s.id) as price_count
        FROM suppliers s ORDER BY s.updated_at DESC
    """).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_supplier(sid: int) -> Optional[Dict]:
    """获取单个供应商"""
    conn = get_db()
    cursor = conn.cursor()
    row = cursor.execute("SELECT * FROM suppliers WHERE id = ?", (sid,)).fetchone()
    conn.close()
    return dict(row) if row else None


def update_supplier(sid: int, **kwargs) -> bool:
    """更新供应商信息"""
    allowed = ['name', 'contact_person', 'phone', 'address', 'materials', 'rating', 'notes']
    updates = {k: v for k, v in kwargs.items() if k in allowed}
    if not updates:
        return False
    updates['updated_at'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    set_clause = ", ".join(f"{k} = ?" for k in updates)
    values = list(updates.values()) + [sid]
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(f"UPDATE suppliers SET {set_clause} WHERE id = ?", values)
    conn.commit()
    conn.close()
    return True


def delete_supplier(sid: int) -> bool:
    """删除供应商"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM supplier_prices WHERE supplier_id = ?", (sid,))
    cursor.execute("DELETE FROM suppliers WHERE id = ?", (sid,))
    conn.commit()
    conn.close()
    return True


# ============================================================
# 供应商报价管理
# ============================================================

def add_supplier_price(supplier_id: int, material_name: str,
                       unit_price: float, price_date: str = "",
                       notes: str = "") -> int:
    """添加供应商报价记录"""
    if not price_date:
        price_date = datetime.now().strftime("%Y-%m-%d")
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO supplier_prices (supplier_id, material_name, unit_price, price_date, notes)
        VALUES (?, ?, ?, ?, ?)
    """, (supplier_id, material_name, unit_price, price_date, notes))
    pid = cursor.lastrowid
    conn.commit()
    conn.close()
    return pid


def get_supplier_prices(supplier_id: int) -> List[Dict]:
    """获取供应商的报价历史"""
    conn = get_db()
    cursor = conn.cursor()
    rows = cursor.execute("""
        SELECT * FROM supplier_prices
        WHERE supplier_id = ?
        ORDER BY price_date DESC, id DESC
    """, (supplier_id,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_best_prices(material_name: str) -> List[Dict]:
    """获取某种材料的最优报价（按价格升序）"""
    conn = get_db()
    cursor = conn.cursor()
    rows = cursor.execute("""
        SELECT sp.*, s.name as supplier_name, s.phone, s.contact_person
        FROM supplier_prices sp
        JOIN suppliers s ON sp.supplier_id = s.id
        WHERE sp.material_name = ?
        ORDER BY sp.unit_price ASC
    """, (material_name,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ============================================================
# 养护记录管理
# ============================================================

def add_curing_record(project_id: int, record_date: str = "",
                      weather: str = "晴", temp_min: float = 20,
                      temp_max: float = 25, humidity: float = 60,
                      wind_level: str = "", curing_measure: str = "",
                      notes: str = "") -> int:
    """添加养护记录"""
    if not record_date:
        record_date = datetime.now().strftime("%Y-%m-%d")
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO curing_records (project_id, record_date, weather,
            temp_min, temp_max, humidity, wind_level, curing_measure, notes)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (project_id, record_date, weather, temp_min, temp_max,
          humidity, wind_level, curing_measure, notes))
    rid = cursor.lastrowid
    conn.commit()
    conn.close()
    return rid


def get_curing_records(project_id: int) -> List[Dict]:
    """获取项目的养护记录"""
    conn = get_db()
    cursor = conn.cursor()
    rows = cursor.execute("""
        SELECT * FROM curing_records
        WHERE project_id = ?
        ORDER BY record_date DESC, id DESC
    """, (project_id,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def delete_curing_record(rid: int) -> bool:
    """删除养护记录"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM curing_records WHERE id = ?", (rid,))
    conn.commit()
    conn.close()
    return True


# ============================================================
# 班组管理
# ============================================================

def add_team(name: str, leader: str = "", specialty: str = "") -> int:
    """添加班组"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("INSERT INTO teams (name, leader, specialty) VALUES (?, ?, ?)",
                   (name, leader, specialty))
    tid = cursor.lastrowid
    conn.commit()
    conn.close()
    return tid


def get_teams() -> List[Dict]:
    """获取所有班组"""
    conn = get_db()
    cursor = conn.cursor()
    rows = cursor.execute("""
        SELECT t.*, (SELECT COUNT(*) FROM workers WHERE team_id = t.id) as actual_members
        FROM teams t ORDER BY t.id
    """).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def delete_team(tid: int) -> bool:
    """删除班组"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE workers SET team_id = 0 WHERE team_id = ?", (tid,))
    cursor.execute("DELETE FROM teams WHERE id = ?", (tid,))
    conn.commit()
    conn.close()
    return True


# ============================================================
# 工人管理
# ============================================================

def add_worker(name: str, phone: str = "", role: str = "工人",
               team_id: int = 0, hourly_rate: float = 0, notes: str = "") -> int:
    """添加工人"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO workers (name, phone, role, team_id, hourly_rate, notes)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (name, phone, role, team_id, hourly_rate, notes))
    wid = cursor.lastrowid
    # 更新班组人数
    if team_id:
        cursor.execute("UPDATE teams SET member_count = (SELECT COUNT(*) FROM workers WHERE team_id = ?) WHERE id = ?",
                       (team_id, team_id))
    conn.commit()
    conn.close()
    return wid


def get_workers(team_id: int = 0) -> List[Dict]:
    """获取工人列表"""
    conn = get_db()
    cursor = conn.cursor()
    if team_id:
        rows = cursor.execute("""
            SELECT w.*, t.name as team_name FROM workers w
            LEFT JOIN teams t ON w.team_id = t.id
            WHERE w.team_id = ? ORDER BY w.name
        """, (team_id,)).fetchall()
    else:
        rows = cursor.execute("""
            SELECT w.*, t.name as team_name FROM workers w
            LEFT JOIN teams t ON w.team_id = t.id
            ORDER BY w.team_id, w.name
        """).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_worker(wid: int) -> Optional[Dict]:
    """获取单个工人"""
    conn = get_db()
    cursor = conn.cursor()
    row = cursor.execute("""
        SELECT w.*, t.name as team_name FROM workers w
        LEFT JOIN teams t ON w.team_id = t.id WHERE w.id = ?
    """, (wid,)).fetchone()
    conn.close()
    return dict(row) if row else None


def update_worker(wid: int, **kwargs) -> bool:
    """更新工人信息"""
    allowed = ['name', 'phone', 'role', 'team_id', 'hourly_rate', 'notes']
    updates = {k: v for k, v in kwargs.items() if k in allowed}
    if not updates:
        return False
    set_clause = ", ".join(f"{k} = ?" for k in updates)
    values = list(updates.values()) + [wid]
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(f"UPDATE workers SET {set_clause} WHERE id = ?", values)
    conn.commit()
    conn.close()
    return True


def delete_worker(wid: int) -> bool:
    """删除工人"""
    conn = get_db()
    cursor = conn.cursor()
    worker = cursor.execute("SELECT team_id FROM workers WHERE id = ?", (wid,)).fetchone()
    cursor.execute("DELETE FROM workers WHERE id = ?", (wid,))
    if worker and worker['team_id']:
        cursor.execute("UPDATE teams SET member_count = (SELECT COUNT(*) FROM workers WHERE team_id = ?) WHERE id = ?",
                       (worker['team_id'], worker['team_id']))
    conn.commit()
    conn.close()
    return True


# ============================================================
# 工时管理
# ============================================================

def add_work_hours(project_id: int, worker_id: int, work_date: str = "",
                   hours: float = 8, work_type: str = "", notes: str = "") -> int:
    """添加工时记录"""
    if not work_date:
        work_date = datetime.now().strftime("%Y-%m-%d")
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO work_hours (project_id, worker_id, work_date, hours, work_type, notes)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (project_id, worker_id, work_date, hours, work_type, notes))
    wid = cursor.lastrowid
    conn.commit()
    conn.close()
    return wid


def get_work_hours(project_id: int) -> List[Dict]:
    """获取项目的工时记录"""
    conn = get_db()
    cursor = conn.cursor()
    rows = cursor.execute("""
        SELECT wh.*, w.name as worker_name, w.role, t.name as team_name
        FROM work_hours wh
        JOIN workers w ON wh.worker_id = w.id
        LEFT JOIN teams t ON w.team_id = t.id
        WHERE wh.project_id = ?
        ORDER BY wh.work_date DESC, wh.id DESC
    """, (project_id,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_work_hours_summary(project_id: int) -> Dict:
    """获取项目工时汇总"""
    conn = get_db()
    cursor = conn.cursor()
    rows = cursor.execute("""
        SELECT wh.work_type, COUNT(DISTINCT wh.worker_id) as worker_count,
               SUM(wh.hours) as total_hours,
               SUM(wh.hours * w.hourly_rate) as total_cost
        FROM work_hours wh
        JOIN workers w ON wh.worker_id = w.id
        WHERE wh.project_id = ?
        GROUP BY wh.work_type
    """, (project_id,)).fetchall()

    total_hours = 0
    total_cost = 0
    details = []
    for r in rows:
        d = dict(r)
        total_hours += d['total_hours'] or 0
        total_cost += d['total_cost'] or 0
        details.append(d)

    conn.close()
    return {"total_hours": round(total_hours, 1), "total_cost": round(total_cost, 2),
            "details": details}


# ============================================================
# 成本预算管理
# ============================================================

def add_budget_item(project_id: int, category: str, planned_amount: float = 0,
                    actual_amount: float = 0, description: str = "") -> int:
    """添加预算项"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO budget_items (project_id, category, planned_amount, actual_amount, description)
        VALUES (?, ?, ?, ?, ?)
    """, (project_id, category, planned_amount, actual_amount, description))
    bid = cursor.lastrowid
    conn.commit()
    conn.close()
    return bid


def get_budget_items(project_id: int) -> List[Dict]:
    """获取项目的预算项"""
    conn = get_db()
    cursor = conn.cursor()
    rows = cursor.execute("""
        SELECT * FROM budget_items WHERE project_id = ? ORDER BY id
    """, (project_id,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_budget_summary(project_id: int) -> Dict:
    """获取项目预算汇总"""
    conn = get_db()
    cursor = conn.cursor()
    rows = cursor.execute("""
        SELECT category,
               SUM(planned_amount) as planned,
               SUM(actual_amount) as actual
        FROM budget_items WHERE project_id = ?
        GROUP BY category ORDER BY category
    """, (project_id,)).fetchall()

    categories = []
    total_planned = 0
    total_actual = 0
    for r in rows:
        d = dict(r)
        categories.append(d)
        total_planned += d['planned'] or 0
        total_actual += d['actual'] or 0

    conn.close()
    return {
        "total_planned": round(total_planned, 2),
        "total_actual": round(total_actual, 2),
        "variance": round(total_planned - total_actual, 2),
        "categories": categories,
    }


def update_budget_item(bid: int, **kwargs) -> bool:
    """更新预算项"""
    allowed = ['planned_amount', 'actual_amount', 'description']
    updates = {k: v for k, v in kwargs.items() if k in allowed}
    if not updates:
        return False
    set_clause = ", ".join(f"{k} = ?" for k in updates)
    values = list(updates.values()) + [bid]
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(f"UPDATE budget_items SET {set_clause} WHERE id = ?", values)
    conn.commit()
    conn.close()
    return True


def delete_budget_item(bid: int) -> bool:
    """删除预算项"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM budget_items WHERE id = ?", (bid,))
    conn.commit()
    conn.close()
    return True


# ============================================================
# 通知提醒系统
# ============================================================

def add_notification(project_id: int, notif_type: str, title: str,
                     message: str = "", due_date: str = "") -> int:
    """添加通知"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO notifications (project_id, type, title, message, due_date)
        VALUES (?, ?, ?, ?, ?)
    """, (project_id, notif_type, title, message, due_date))
    nid = cursor.lastrowid
    conn.commit()
    conn.close()
    return nid


def get_notifications(project_id: int = 0, unread_only: bool = False) -> List[Dict]:
    """获取通知列表"""
    conn = get_db()
    cursor = conn.cursor()
    if project_id:
        if unread_only:
            rows = cursor.execute("""
                SELECT n.*, p.name as project_name FROM notifications n
                JOIN projects p ON n.project_id = p.id
                WHERE n.project_id = ? AND n.is_read = 0
                ORDER BY n.created_at DESC
            """, (project_id,)).fetchall()
        else:
            rows = cursor.execute("""
                SELECT n.*, p.name as project_name FROM notifications n
                JOIN projects p ON n.project_id = p.id
                WHERE n.project_id = ?
                ORDER BY n.created_at DESC LIMIT 50
            """, (project_id,)).fetchall()
    else:
        if unread_only:
            rows = cursor.execute("""
                SELECT n.*, p.name as project_name FROM notifications n
                JOIN projects p ON n.project_id = p.id
                WHERE n.is_read = 0 ORDER BY n.created_at DESC
            """).fetchall()
        else:
            rows = cursor.execute("""
                SELECT n.*, p.name as project_name FROM notifications n
                JOIN projects p ON n.project_id = p.id
                ORDER BY n.created_at DESC LIMIT 100
            """).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def mark_notification_read(nid: int) -> bool:
    """标记通知为已读"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE notifications SET is_read = 1 WHERE id = ?", (nid,))
    conn.commit()
    conn.close()
    return True


def mark_all_notifications_read(project_id: int = 0) -> int:
    """标记所有通知为已读"""
    conn = get_db()
    cursor = conn.cursor()
    if project_id:
        cursor.execute("UPDATE notifications SET is_read = 1 WHERE project_id = ? AND is_read = 0",
                       (project_id,))
    else:
        cursor.execute("UPDATE notifications SET is_read = 1 WHERE is_read = 0")
    count = cursor.rowcount
    conn.commit()
    conn.close()
    return count


def delete_notification(nid: int) -> bool:
    """删除通知"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM notifications WHERE id = ?", (nid,))
    conn.commit()
    conn.close()
    return True


def create_curing_reminder(project_id: int) -> int:
    """自动创建养护提醒"""
    project = get_project(project_id)
    if not project:
        return 0
    title = f"🔔 {project['name']} - 养护提醒"
    message = f"项目「{project['name']}」已达到养护阶段，请检查养护措施并记录养护日志。"
    return add_notification(project_id, "curing_reminder", title, message)


def create_quality_test_reminder(project_id: int) -> int:
    """自动创建质量检测提醒"""
    project = get_project(project_id)
    if not project:
        return 0
    title = f"🧪 {project['name']} - 质量检测提醒"
    message = f"项目「{project['name']}」需进行质量检测，请安排检测并记录结果。"
    return add_notification(project_id, "test_reminder", title, message)


# ============================================================
# 设备管理
# ============================================================

def add_equipment(name: str, type_: str = "", model: str = "",
                  quantity: int = 1, unit: str = "台", status: str = "可用",
                  purchase_date: str = "", next_maintenance: str = "",
                  notes: str = "") -> int:
    """添加设备"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO equipment (name, type, model, quantity, unit, status,
                               purchase_date, next_maintenance, notes)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (name, type_, model, quantity, unit, status,
          purchase_date, next_maintenance, notes))
    eid = cursor.lastrowid
    conn.commit()
    conn.close()
    return eid


def get_equipment(eid: int = 0) -> List[Dict]:
    """获取设备列表"""
    conn = get_db()
    cursor = conn.cursor()
    if eid:
        row = cursor.execute("SELECT * FROM equipment WHERE id = ?", (eid,)).fetchone()
        conn.close()
        return dict(row) if row else None
    rows = cursor.execute("""
        SELECT e.*, 
            (SELECT COUNT(*) FROM equipment_usage WHERE equipment_id = e.id) as usage_count
        FROM equipment e ORDER BY e.type, e.name
    """).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def update_equipment(eid: int, **kwargs) -> bool:
    """更新设备信息"""
    allowed = ['name', 'type', 'model', 'quantity', 'unit', 'status',
               'purchase_date', 'next_maintenance', 'notes']
    updates = {k: v for k, v in kwargs.items() if k in allowed}
    if not updates:
        return False
    set_clause = ", ".join(f"{k} = ?" for k in updates)
    values = list(updates.values()) + [eid]
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(f"UPDATE equipment SET {set_clause} WHERE id = ?", values)
    conn.commit()
    conn.close()
    return True


def delete_equipment(eid: int) -> bool:
    """删除设备"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM equipment_usage WHERE equipment_id = ?", (eid,))
    cursor.execute("DELETE FROM equipment WHERE id = ?", (eid,))
    conn.commit()
    conn.close()
    return True


def add_equipment_usage(equipment_id: int, project_id: int,
                        quantity_used: int = 1, start_date: str = "",
                        end_date: str = "", operator: str = "",
                        notes: str = "") -> int:
    """记录设备使用"""
    if not start_date:
        start_date = datetime.now().strftime("%Y-%m-%d")
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO equipment_usage (equipment_id, project_id, quantity_used,
                                     start_date, end_date, operator, notes)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (equipment_id, project_id, quantity_used, start_date, end_date, operator, notes))
    uid = cursor.lastrowid
    conn.commit()
    conn.close()
    return uid


def get_equipment_usage(project_id: int = 0) -> List[Dict]:
    """获取设备使用记录"""
    conn = get_db()
    cursor = conn.cursor()
    if project_id:
        rows = cursor.execute("""
            SELECT u.*, e.name as equipment_name, e.type as equipment_type,
                   p.name as project_name
            FROM equipment_usage u
            JOIN equipment e ON u.equipment_id = e.id
            JOIN projects p ON u.project_id = p.id
            WHERE u.project_id = ?
            ORDER BY u.start_date DESC
        """, (project_id,)).fetchall()
    else:
        rows = cursor.execute("""
            SELECT u.*, e.name as equipment_name, e.type as equipment_type,
                   p.name as project_name
            FROM equipment_usage u
            JOIN equipment e ON u.equipment_id = e.id
            JOIN projects p ON u.project_id = p.id
            ORDER BY u.start_date DESC LIMIT 100
        """).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ============================================================
# 验收管理
# ============================================================

def add_acceptance(project_id: int, acceptance_type: str,
                   check_date: str = "", inspector: str = "",
                   result: str = "待定", defects: str = "",
                   score: float = 0, notes: str = "") -> int:
    """创建验收记录"""
    if not check_date:
        check_date = datetime.now().strftime("%Y-%m-%d")
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO acceptance_records (project_id, acceptance_type, check_date,
            inspector, result, defects, score, notes)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (project_id, acceptance_type, check_date, inspector, result, defects, score, notes))
    aid = cursor.lastrowid

    # 自动创建验收项目（根据验收类型）
    items = _get_acceptance_default_items(acceptance_type)
    for item in items:
        cursor.execute("""
            INSERT INTO acceptance_items (acceptance_id, item_name, standard)
            VALUES (?, ?, ?)
        """, (aid, item['name'], item['standard']))

    conn.commit()
    conn.close()
    return aid


def _get_acceptance_default_items(acceptance_type: str) -> List[Dict]:
    """根据验收类型返回默认检查项"""
    templates = {
        "基层验收": [
            {"name": "基层平整度", "standard": "2m靠尺≤3mm"},
            {"name": "基层强度", "standard": "≥C20, 表面无起砂"},
            {"name": "基层清洁度", "standard": "无油污、浮灰、松散颗粒"},
            {"name": "基层含水率", "standard": "≤8%"},
            {"name": "伸缩缝设置", "standard": "按设计要求留设"},
        ],
        "中间验收": [
            {"name": "抗裂砂浆厚度", "standard": "设计厚度±2mm"},
            {"name": "钢纤维分布", "standard": "均匀无结团"},
            {"name": "表面平整度", "standard": "2m靠尺≤3mm"},
            {"name": "养护时间", "standard": "≥7天"},
            {"name": "界面处理", "standard": "涂刷均匀无漏涂"},
        ],
        "竣工验收": [
            {"name": "面层平整度", "standard": "2m靠尺≤2mm"},
            {"name": "面层光泽度", "standard": "≥60°"},
            {"name": "颜色均匀性", "standard": "无明显色差"},
            {"name": "硬度测试", "standard": "莫氏硬度≥6"},
            {"name": "抗渗性能", "standard": "24h无渗水"},
            {"name": "表面缺陷", "standard": "无裂纹、空鼓、麻面"},
        ],
    }
    return templates.get(acceptance_type, [
        {"name": "外观质量", "standard": "符合规范要求"},
        {"name": "尺寸偏差", "standard": "在允许范围内"},
    ])


def get_acceptances(project_id: int) -> List[Dict]:
    """获取项目的验收记录"""
    conn = get_db()
    cursor = conn.cursor()
    rows = cursor.execute("""
        SELECT * FROM acceptance_records WHERE project_id = ?
        ORDER BY check_date DESC, id DESC
    """, (project_id,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_acceptance(aid: int) -> Optional[Dict]:
    """获取验收详情（含检查项）"""
    conn = get_db()
    cursor = conn.cursor()
    row = cursor.execute("SELECT * FROM acceptance_records WHERE id = ?", (aid,)).fetchone()
    if not row:
        conn.close()
        return None
    acceptance = dict(row)
    items = cursor.execute("""
        SELECT * FROM acceptance_items WHERE acceptance_id = ? ORDER BY id
    """, (aid,)).fetchall()
    acceptance['items'] = [dict(r) for r in items]
    conn.close()
    return acceptance


def update_acceptance_item(item_id: int, is_pass: int, actual_value: str = "") -> bool:
    """更新验收检查项"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE acceptance_items SET is_pass = ?, actual_value = ? WHERE id = ?
    """, (is_pass, actual_value, item_id))
    conn.commit()
    conn.close()
    return True


def update_acceptance(aid: int, **kwargs) -> bool:
    """更新验收记录"""
    allowed = ['result', 'defects', 'score', 'notes', 'inspector']
    updates = {k: v for k, v in kwargs.items() if k in allowed}
    if not updates:
        return False
    set_clause = ", ".join(f"{k} = ?" for k in updates)
    values = list(updates.values()) + [aid]
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(f"UPDATE acceptance_records SET {set_clause} WHERE id = ?", values)
    conn.commit()
    conn.close()
    return True


def delete_acceptance(aid: int) -> bool:
    """删除验收记录"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM acceptance_items WHERE acceptance_id = ?", (aid,))
    cursor.execute("DELETE FROM acceptance_records WHERE id = ?", (aid,))
    conn.commit()
    conn.close()
    return True


# ============================================================
# 安全管理
# ============================================================

SAFETY_CHECK_TEMPLATES = {
    "日常检查": [
        {"item": "安全帽佩戴", "standard": "全员正确佩戴"},
        {"item": "高处作业防护", "standard": "安全带牢固、脚手架稳固"},
        {"item": "用电安全", "standard": "电缆无破损、漏保有效"},
        {"item": "消防器材", "standard": "灭火器在位、压力正常"},
        {"item": "施工通道", "standard": "畅通无杂物堆放"},
        {"item": "材料堆放", "standard": "整齐有序、不超高"},
        {"item": "机械操作", "standard": "持证上岗、防护装置完好"},
    ],
    "专项检查": [
        {"item": "临时用电系统", "standard": "三级配电、两级保护"},
        {"item": "动火作业审批", "standard": "审批手续齐全、监护人在场"},
        {"item": "吊装作业", "standard": "吊具完好、指挥信号明确"},
        {"item": "危化品管理", "standard": "专人保管、存放合规"},
        {"item": "应急预案", "standard": "预案完善、演练记录齐全"},
    ],
    "周检": [
        {"item": "现场文明施工", "standard": "工完料清、场地整洁"},
        {"item": "防护设施", "standard": "临边洞口防护到位"},
        {"item": "施工机具", "standard": "完好无损、定期保养"},
        {"item": "安全教育", "standard": "每日班前会议记录"},
        {"item": "劳保用品", "standard": "配备齐全、正确使用"},
    ],
}


def add_safety_check(project_id: int, check_date: str = "",
                     inspector: str = "", check_type: str = "日常检查",
                     items: str = "[]", result: str = "合格",
                     rectification: str = "", notes: str = "") -> int:
    """添加安全检查记录"""
    if not check_date:
        check_date = datetime.now().strftime("%Y-%m-%d")
    conn = get_db()
    cursor = conn.cursor()

    # 解析items统计
    import json as _json
    try:
        item_list = _json.loads(items) if isinstance(items, str) else items
    except:
        item_list = []
    total = len(item_list)
    passed = sum(1 for i in item_list if i.get('passed', False))

    cursor.execute("""
        INSERT INTO safety_checks (project_id, check_date, inspector, check_type,
            items, total_items, passed_items, result, rectification, notes)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (project_id, check_date, inspector, check_type, items,
          total, passed, result, rectification, notes))
    sid = cursor.lastrowid
    conn.commit()
    conn.close()
    return sid


def get_safety_checks(project_id: int) -> List[Dict]:
    """获取项目的安全检查记录"""
    conn = get_db()
    cursor = conn.cursor()
    rows = cursor.execute("""
        SELECT * FROM safety_checks WHERE project_id = ?
        ORDER BY check_date DESC, id DESC
    """, (project_id,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_safety_check(sid: int) -> Optional[Dict]:
    """获取安全检查详情"""
    conn = get_db()
    cursor = conn.cursor()
    row = cursor.execute("SELECT * FROM safety_checks WHERE id = ?", (sid,)).fetchone()
    conn.close()
    return dict(row) if row else None


def delete_safety_check(sid: int) -> bool:
    """删除安全检查记录"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM safety_checks WHERE id = ?", (sid,))
    conn.commit()
    conn.close()
    return True


def get_safety_check_templates() -> Dict:
    """获取安全检查模板"""
    return SAFETY_CHECK_TEMPLATES


# ============================================================
# 安全事故管理
# ============================================================

def add_safety_incident(project_id: int, incident_date: str = "",
                        incident_type: str = "其他", severity: str = "轻微",
                        description: str = "", injured_person: str = "",
                        treatment: str = "", root_cause: str = "",
                        preventive_measures: str = "") -> int:
    """记录安全事故"""
    if not incident_date:
        incident_date = datetime.now().strftime("%Y-%m-%d")
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO safety_incidents (project_id, incident_date, incident_type,
            severity, description, injured_person, treatment, root_cause,
            preventive_measures)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (project_id, incident_date, incident_type, severity,
          description, injured_person, treatment, root_cause, preventive_measures))
    iid = cursor.lastrowid

    # 在同一连接中创建通知（避免循环导入）
    title = f"🚨 安全事故报告 - {incident_type}"
    message = f"{severity}事故: {(description or '')[:100]}"
    cursor.execute("""
        INSERT INTO notifications (project_id, type, title, message, due_date)
        VALUES (?, ?, ?, ?, ?)
    """, (project_id, "safety_alert", title, message, incident_date))

    conn.commit()
    conn.close()
    return iid


def get_safety_incidents(project_id: int) -> List[Dict]:
    """获取项目的安全事故记录"""
    conn = get_db()
    cursor = conn.cursor()
    rows = cursor.execute("""
        SELECT * FROM safety_incidents WHERE project_id = ?
        ORDER BY incident_date DESC, id DESC
    """, (project_id,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def update_safety_incident(iid: int, **kwargs) -> bool:
    """更新安全事故记录"""
    allowed = ['is_closed', 'closed_date', 'treatment', 'root_cause',
               'preventive_measures', 'severity', 'description']
    updates = {k: v for k, v in kwargs.items() if k in allowed}
    if not updates:
        return False
    set_clause = ", ".join(f"{k} = ?" for k in updates)
    values = list(updates.values()) + [iid]
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(f"UPDATE safety_incidents SET {set_clause} WHERE id = ?", values)
    conn.commit()
    conn.close()
    return True


def delete_safety_incident(iid: int) -> bool:
    """删除安全事故记录"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM safety_incidents WHERE id = ?", (iid,))
    conn.commit()
    conn.close()
    return True


# ============================================================
# 文档管理
# ============================================================

DOCUMENTS_DIR = os.path.join(os.path.dirname(__file__), 'data', 'documents')


def _ensure_doc_dir():
    """确保文档目录存在"""
    os.makedirs(DOCUMENTS_DIR, exist_ok=True)


def add_document(project_id: int, title: str, doc_type: str = "other",
                 description: str = "", filename: str = "",
                 file_data: bytes = b"", version: str = "v1.0",
                 tags: str = "[]") -> Dict:
    """
    添加文档

    参数:
        project_id: 项目ID (0=通用文档)
        title: 文档标题
        doc_type: 文档类型
        description: 描述
        filename: 文件名
        file_data: 文件二进制数据
        version: 版本号
        tags: 标签JSON数组

    返回:
        文档记录
    """
    _ensure_doc_dir()

    # 保存文件
    import uuid
    safe_name = f"{uuid.uuid4().hex}_{filename}" if filename else f"{uuid.uuid4().hex}.bin"
    filepath = os.path.join(DOCUMENTS_DIR, safe_name)
    with open(filepath, 'wb') as f:
        f.write(file_data)
    file_size = os.path.getsize(filepath)

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO documents (project_id, doc_type, title, description,
            filename, file_size, version, tags)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (project_id, doc_type, title, description, safe_name,
          file_size, version, tags))
    did = cursor.lastrowid
    conn.commit()
    conn.close()

    return {
        "id": did,
        "title": title,
        "filename": safe_name,
        "file_size": file_size,
        "url": f"/api/documents/{did}/download",
    }


def get_documents(project_id: int = 0, doc_type: str = "") -> List[Dict]:
    """获取文档列表"""
    conn = get_db()
    cursor = conn.cursor()
    if project_id and doc_type:
        rows = cursor.execute("""
            SELECT d.*, p.name as project_name FROM documents d
            LEFT JOIN projects p ON d.project_id = p.id
            WHERE d.project_id = ? AND d.doc_type = ?
            ORDER BY d.created_at DESC
        """, (project_id, doc_type)).fetchall()
    elif project_id:
        rows = cursor.execute("""
            SELECT d.*, p.name as project_name FROM documents d
            LEFT JOIN projects p ON d.project_id = p.id
            WHERE d.project_id = ?
            ORDER BY d.created_at DESC
        """, (project_id,)).fetchall()
    else:
        rows = cursor.execute("""
            SELECT d.*, p.name as project_name FROM documents d
            LEFT JOIN projects p ON d.project_id = p.id
            ORDER BY d.created_at DESC LIMIT 100
        """).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_document(did: int) -> Optional[Dict]:
    """获取单个文档"""
    conn = get_db()
    cursor = conn.cursor()
    row = cursor.execute("""
        SELECT d.*, p.name as project_name FROM documents d
        LEFT JOIN projects p ON d.project_id = p.id
        WHERE d.id = ?
    """, (did,)).fetchone()
    conn.close()
    return dict(row) if row else None


def delete_document(did: int) -> bool:
    """删除文档"""
    doc = get_document(did)
    if not doc:
        return False
    filepath = os.path.join(DOCUMENTS_DIR, doc['filename'])
    if os.path.exists(filepath):
        os.remove(filepath)
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM documents WHERE id = ?", (did,))
    conn.commit()
    conn.close()
    return True


# ============================================================
# 材料申购管理
# ============================================================

def add_material_request(project_id: int, applicant: str = "",
                         request_date: str = "", notes: str = "",
                         items: list = None) -> int:
    """创建材料申购单"""
    if not request_date:
        request_date = datetime.now().strftime("%Y-%m-%d")
    if items is None:
        items = []
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO material_requests (project_id, request_date, applicant, notes)
        VALUES (?, ?, ?, ?)
    """, (project_id, request_date, applicant, notes))
    rid = cursor.lastrowid
    for item in items:
        cursor.execute("""
            INSERT INTO request_items (request_id, material_name, quantity_kg, unit, purpose)
            VALUES (?, ?, ?, ?, ?)
        """, (rid, item.get('material_name', ''), float(item.get('quantity_kg', 0)),
              item.get('unit', 'kg'), item.get('purpose', '')))
    conn.commit()
    conn.close()
    return rid


def get_material_requests(project_id: int = 0) -> List[Dict]:
    """获取材料申购单"""
    conn = get_db()
    cursor = conn.cursor()
    if project_id:
        rows = cursor.execute("""
            SELECT r.*, p.name as project_name,
                (SELECT COUNT(*) FROM request_items WHERE request_id = r.id) as item_count
            FROM material_requests r
            JOIN projects p ON r.project_id = p.id
            WHERE r.project_id = ?
            ORDER BY r.created_at DESC
        """, (project_id,)).fetchall()
    else:
        rows = cursor.execute("""
            SELECT r.*, p.name as project_name,
                (SELECT COUNT(*) FROM request_items WHERE request_id = r.id) as item_count
            FROM material_requests r
            JOIN projects p ON r.project_id = p.id
            ORDER BY r.created_at DESC LIMIT 100
        """).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_material_request(rid: int) -> Optional[Dict]:
    """获取申购单详情（含明细）"""
    conn = get_db()
    cursor = conn.cursor()
    row = cursor.execute("""
        SELECT r.*, p.name as project_name FROM material_requests r
        JOIN projects p ON r.project_id = p.id WHERE r.id = ?
    """, (rid,)).fetchone()
    if not row:
        conn.close()
        return None
    req = dict(row)
    items = cursor.execute("""
        SELECT * FROM request_items WHERE request_id = ? ORDER BY id
    """, (rid,)).fetchall()
    req['items'] = [dict(r) for r in items]
    conn.close()
    return req


def update_material_request(rid: int, **kwargs) -> bool:
    """更新申购单状态"""
    allowed = ['status', 'approver', 'approved_date', 'notes']
    updates = {k: v for k, v in kwargs.items() if k in allowed}
    if not updates:
        return False
    set_clause = ", ".join(f"{k} = ?" for k in updates)
    values = list(updates.values()) + [rid]
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(f"UPDATE material_requests SET {set_clause} WHERE id = ?", values)
    # 如果批准，自动创建通知
    if updates.get('status') == '已批准':
        req = cursor.execute(
            "SELECT project_id FROM material_requests WHERE id = ?", (rid,)
        ).fetchone()
        if req:
            cursor.execute("""
                INSERT INTO notifications (project_id, type, title, message)
                VALUES (?, ?, ?, ?)
            """, (req['project_id'], 'material_approved',
                  '✅ 材料申购已批准',
                  f'申购单 #{rid} 已获批准，请安排采购。'))
    conn.commit()
    conn.close()
    return True


def delete_material_request(rid: int) -> bool:
    """删除申购单"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM request_items WHERE request_id = ?", (rid,))
    cursor.execute("DELETE FROM material_requests WHERE id = ?", (rid,))
    conn.commit()
    conn.close()
    return True


# ============================================================
# 分包商管理
# ============================================================

def add_subcontractor(name: str, contact_person: str = "", phone: str = "",
                      scope: str = "", contract_amount: float = 0,
                      contract_date: str = "", start_date: str = "",
                      end_date: str = "", status: str = "进行中",
                      rating: int = 3, notes: str = "") -> int:
    """添加分包商"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO subcontractors (name, contact_person, phone, scope,
            contract_amount, contract_date, start_date, end_date, status, rating, notes)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (name, contact_person, phone, scope, contract_amount,
          contract_date, start_date, end_date, status, rating, notes))
    sid = cursor.lastrowid
    conn.commit()
    conn.close()
    return sid


def get_subcontractors() -> List[Dict]:
    """获取所有分包商"""
    conn = get_db()
    cursor = conn.cursor()
    rows = cursor.execute("""
        SELECT * FROM subcontractors ORDER BY status, name
    """).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def update_subcontractor(sid: int, **kwargs) -> bool:
    """更新分包商"""
    allowed = ['name', 'contact_person', 'phone', 'scope', 'contract_amount',
               'contract_date', 'start_date', 'end_date', 'status', 'rating', 'notes']
    updates = {k: v for k, v in kwargs.items() if k in allowed}
    if not updates:
        return False
    set_clause = ", ".join(f"{k} = ?" for k in updates)
    values = list(updates.values()) + [sid]
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(f"UPDATE subcontractors SET {set_clause} WHERE id = ?", values)
    conn.commit()
    conn.close()
    return True


def delete_subcontractor(sid: int) -> bool:
    """删除分包商"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM subcontractors WHERE id = ?", (sid,))
    conn.commit()
    conn.close()
    return True


# ============================================================
# 项目模板管理
# ============================================================

def add_project_template(name: str, description: str = "",
                         config: dict = None) -> int:
    """保存项目模板"""
    import json as _json
    if config is None:
        config = {}
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO project_templates (name, description, config)
        VALUES (?, ?, ?)
    """, (name, description, _json.dumps(config, ensure_ascii=False)))
    tid = cursor.lastrowid
    conn.commit()
    conn.close()
    return tid


def get_project_templates() -> List[Dict]:
    """获取所有项目模板"""
    conn = get_db()
    cursor = conn.cursor()
    rows = cursor.execute("""
        SELECT * FROM project_templates ORDER BY created_at DESC
    """).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def delete_project_template(tid: int) -> bool:
    """删除项目模板"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM project_templates WHERE id = ?", (tid,))
    conn.commit()
    conn.close()
    return True


# ============================================================
# 材料用量对比
# ============================================================

def set_material_consumption(project_id: int, material_name: str,
                              planned_kg: float = 0, actual_kg: float = 0,
                              unit: str = "kg", notes: str = "") -> int:
    """设置材料用量（计划vs实际）"""
    conn = get_db()
    cursor = conn.cursor()
    # 检查是否已存在
    existing = cursor.execute("""
        SELECT id FROM material_consumption
        WHERE project_id = ? AND material_name = ?
    """, (project_id, material_name)).fetchone()
    if existing:
        cursor.execute("""
            UPDATE material_consumption SET planned_kg = ?, actual_kg = ?,
                unit = ?, notes = ?
            WHERE id = ?
        """, (planned_kg, actual_kg, unit, notes, existing['id']))
        cid = existing['id']
    else:
        cursor.execute("""
            INSERT INTO material_consumption (project_id, material_name, planned_kg, actual_kg, unit, notes)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (project_id, material_name, planned_kg, actual_kg, unit, notes))
        cid = cursor.lastrowid
    conn.commit()
    conn.close()
    return cid


def get_material_consumption(project_id: int) -> List[Dict]:
    """获取项目材料用量对比"""
    conn = get_db()
    cursor = conn.cursor()
    rows = cursor.execute("""
        SELECT * FROM material_consumption WHERE project_id = ?
        ORDER BY material_name
    """, (project_id,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ============================================================
# 日报生成
# ============================================================

def generate_daily_report(project_id: int, report_date: str = "") -> Dict:
    """
    生成施工日报

    参数:
        project_id: 项目ID
        report_date: 日期（默认今天）

    返回:
        日报数据
    """
    if not report_date:
        from datetime import date
        report_date = date.today().strftime("%Y-%m-%d")

    project = get_project(project_id)
    if not project:
        return {"error": "项目不存在"}

    conn = get_db()
    cursor = conn.cursor()

    # 当天日志
    log = cursor.execute("""
        SELECT * FROM daily_logs WHERE project_id = ? AND log_date = ?
    """, (project_id, report_date)).fetchone()
    daily_log = dict(log) if log else None

    # 当天工时
    work_hours_rows = cursor.execute("""
        SELECT w.name, w.role, wh.hours, wh.work_type
        FROM work_hours wh
        JOIN workers w ON wh.worker_id = w.id
        WHERE wh.project_id = ? AND wh.work_date = ?
    """, (project_id, report_date)).fetchall()
    today_hours = [dict(r) for r in work_hours_rows]
    total_workers = len(set(r['name'] for r in today_hours))
    total_hours = sum(r['hours'] or 0 for r in today_hours)

    # 当天材料记录
    materials_today = cursor.execute("""
        SELECT material_name, quantity_kg, unit_price, total_cost
        FROM material_records WHERE project_id = ? AND record_date = ?
    """, (project_id, report_date)).fetchall()

    # 当天养护记录
    curing_today = cursor.execute("""
        SELECT * FROM curing_records WHERE project_id = ? AND record_date = ?
    """, (project_id, report_date)).fetchall()

    # 当天检查进度
    checklist = cursor.execute("""
        SELECT COUNT(*) as total,
               SUM(CASE WHEN is_checked=1 THEN 1 ELSE 0 END) as checked
        FROM checklist_state WHERE project_id = ?
    """, (project_id,)).fetchone()

    # 当天验收
    acc_today = cursor.execute("""
        SELECT * FROM acceptance_records
        WHERE project_id = ? AND check_date = ?
    """, (project_id, report_date)).fetchall()

    conn.close()

    # 累计进度
    dashboard = get_dashboard_data(project_id)

    return {
        "project_name": project['name'],
        "report_date": report_date,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "weather": daily_log['weather'] if daily_log else '未记录',
        "temperature": daily_log['temperature'] if daily_log else '',
        "daily_log": daily_log,
        "work_hours": {
            "total_workers": total_workers,
            "total_hours": total_hours,
            "details": today_hours,
        },
        "materials": [dict(r) for r in materials_today],
        "curing": [dict(r) for r in curing_today],
        "acceptance": [dict(r) for r in acc_today],
        "checklist_progress": {
            "total": checklist['total'] or 0,
            "checked": checklist['checked'] or 0,
            "percent": round((checklist['checked'] or 0) / max(checklist['total'] or 1, 1) * 100, 1),
        },
        "overall_progress": dashboard.get('overall_progress', 0),
    }


# ============================================================
# 全局搜索
# ============================================================

def global_search(query: str) -> Dict:
    """
    全局搜索

    参数:
        query: 搜索关键词

    返回:
        按分类的结果
    """
    conn = get_db()
    cursor = conn.cursor()
    q = f"%{query}%"

    results = {}

    # 项目
    projects = cursor.execute("""
        SELECT id, name, 'project' as type, status as info
        FROM projects WHERE name LIKE ? OR location LIKE ? OR notes LIKE ?
        LIMIT 10
    """, (q, q, q)).fetchall()
    results['projects'] = [dict(r) for r in projects]

    # 施工日志
    logs = cursor.execute("""
        SELECT ld.id, ld.work_content as name, 'log' as type, p.name as info
        FROM daily_logs ld JOIN projects p ON ld.project_id = p.id
        WHERE ld.work_content LIKE ? OR ld.issues LIKE ? OR ld.materials_used LIKE ?
        LIMIT 10
    """, (q, q, q)).fetchall()
    results['logs'] = [dict(r) for r in logs]

    # 供应商
    suppliers = cursor.execute("""
        SELECT id, name, 'supplier' as type, CONCAT(contact_person, ' ', phone) as info
        FROM suppliers WHERE name LIKE ? OR contact_person LIKE ? OR phone LIKE ?
        LIMIT 10
    """, (q, q, q)).fetchall()
    results['suppliers'] = [dict(r) for r in suppliers]

    # 工人
    workers = cursor.execute("""
        SELECT id, name, 'worker' as type, role as info
        FROM workers WHERE name LIKE ? OR phone LIKE ? OR role LIKE ?
        LIMIT 10
    """, (q, q, q)).fetchall()
    results['workers'] = [dict(r) for r in workers]

    # 设备
    eq = cursor.execute("""
        SELECT id, name, 'equipment' as type, CONCAT(type, ' ', status) as info
        FROM equipment WHERE name LIKE ? OR type LIKE ? OR model LIKE ?
        LIMIT 10
    """, (q, q, q)).fetchall()
    results['equipment'] = [dict(r) for r in eq]

    conn.close()
    return results


# ============================================================
# 项目分析看板
# ============================================================

def get_analytics() -> Dict:
    """跨项目分析数据"""
    conn = get_db()
    cursor = conn.cursor()

    # 项目统计
    total_projects = cursor.execute("SELECT COUNT(*) FROM projects").fetchone()[0]
    active_projects = cursor.execute("SELECT COUNT(*) FROM projects WHERE status='进行中'").fetchone()[0]
    completed_projects = cursor.execute("SELECT COUNT(*) FROM projects WHERE status='已完成'").fetchone()[0]

    total_area = cursor.execute("SELECT COALESCE(SUM(area),0) FROM projects").fetchone()[0]

    # 检查清单总体进度
    checklist = cursor.execute("""
        SELECT COUNT(*) as total, SUM(CASE WHEN is_checked=1 THEN 1 ELSE 0 END) as checked
        FROM checklist_state
    """).fetchone()
    check_total = checklist['total'] or 0
    check_done = checklist['checked'] or 0

    # 材料总用量
    total_material_kg = cursor.execute("""
        SELECT COALESCE(SUM(quantity_kg),0) FROM material_records
    """).fetchone()[0]

    # 总预算
    budget = cursor.execute("""
        SELECT COALESCE(SUM(planned_amount),0) as planned,
               COALESCE(SUM(actual_amount),0) as actual
        FROM budget_items
    """).fetchone()

    # 总工时
    total_hours = cursor.execute("""
        SELECT COALESCE(SUM(hours),0) FROM work_hours
    """).fetchone()[0]

    # 最近施工日志
    recent_logs = [dict(r) for r in cursor.execute("""
        SELECT ld.*, p.name as project_name FROM daily_logs ld
        JOIN projects p ON ld.project_id = p.id
        ORDER BY ld.created_at DESC LIMIT 10
    """).fetchall()]

    # 待审批申购
    pending_requests = cursor.execute("""
        SELECT COUNT(*) FROM material_requests WHERE status='待审批'
    """).fetchone()[0]

    conn.close()

    return {
        "projects": {"total": total_projects, "active": active_projects,
                      "completed": completed_projects, "total_area_m2": total_area},
        "checklist": {"total": check_total, "completed": check_done,
                       "progress": round(check_done/check_total*100,1) if check_total > 0 else 0},
        "materials": {"total_kg": total_material_kg},
        "budget": {"planned": round(budget['planned'],2),
                   "actual": round(budget['actual'],2)},
        "labor": {"total_hours": round(total_hours,1)},
        "pending_requests": pending_requests,
        "recent_logs": recent_logs,
    }


# ============================================================
# CSV/Excel 导出
# ============================================================

def export_to_csv(project_id: int, table_name: str) -> str:
    """
    将项目数据导出为CSV格式

    参数:
        project_id: 项目ID (0=所有项目)
        table_name: 表名 (logs, quality, materials, budget, work_hours, curing)

    返回:
        CSV字符串
    """
    conn = get_db()
    cursor = conn.cursor()

    if table_name == 'logs':
        if project_id:
            rows = cursor.execute("""
                SELECT ld.log_date as 日期, ld.weather as 天气, ld.temperature as 温度,
                       ld.workers as 工人数, ld.work_content as 施工内容,
                       ld.materials_used as 材料使用, ld.issues as 问题
                FROM daily_logs ld WHERE ld.project_id = ? ORDER BY ld.log_date
            """, (project_id,)).fetchall()
        else:
            rows = cursor.execute("""
                SELECT p.name as 项目, ld.log_date as 日期, ld.weather as 天气,
                       ld.temperature as 温度, ld.workers as 工人数,
                       ld.work_content as 施工内容, ld.issues as 问题
                FROM daily_logs ld JOIN projects p ON ld.project_id = p.id
                ORDER BY ld.log_date
            """).fetchall()
    elif table_name == 'materials':
        if project_id:
            rows = cursor.execute("""
                SELECT mr.material_name as 材料名称, mr.quantity_kg as 数量kg,
                       mr.quantity_packages as 包装数, mr.unit_price as 单价,
                       mr.total_cost as 总价, mr.record_date as 记录日期
                FROM material_records mr WHERE mr.project_id = ? ORDER BY mr.record_date
            """, (project_id,)).fetchall()
        else:
            rows = cursor.execute("""
                SELECT p.name as 项目, mr.material_name as 材料名称, mr.quantity_kg as 数量kg,
                       mr.unit_price as 单价, mr.total_cost as 总价, mr.record_date as 记录日期
                FROM material_records mr JOIN projects p ON mr.project_id = p.id
                ORDER BY mr.record_date
            """).fetchall()
    elif table_name == 'quality':
        if project_id:
            rows = cursor.execute("""
                SELECT qt.test_name as 检测项目, qt.standard_value as 标准值,
                       qt.actual_value as 实测值, qt.is_pass as 是否合格, qt.test_date as 检测日期
                FROM quality_tests qt WHERE qt.project_id = ? ORDER BY qt.test_date
            """, (project_id,)).fetchall()
        else:
            rows = cursor.execute("""
                SELECT p.name as 项目, qt.test_name as 检测项目, qt.standard_value as 标准值,
                       qt.actual_value as 实测值, qt.is_pass as 是否合格, qt.test_date as 检测日期
                FROM quality_tests qt JOIN projects p ON qt.project_id = p.id
                ORDER BY qt.test_date
            """).fetchall()
    elif table_name == 'budget':
        if project_id:
            rows = cursor.execute("""
                SELECT category as 类别, planned_amount as 预算金额,
                       actual_amount as 实际支出, description as 说明
                FROM budget_items WHERE project_id = ? ORDER BY category
            """, (project_id,)).fetchall()
        else:
            rows = cursor.execute("""
                SELECT p.name as 项目, category as 类别, planned_amount as 预算金额,
                       actual_amount as 实际支出
                FROM budget_items JOIN projects p ON budget_items.project_id = p.id
                ORDER BY p.name, category
            """).fetchall()
    elif table_name == 'work_hours':
        if project_id:
            rows = cursor.execute("""
                SELECT wh.work_date as 日期, w.name as 工人, w.role as 角色,
                       wh.hours as 工时, wh.work_type as 施工类型
                FROM work_hours wh JOIN workers w ON wh.worker_id = w.id
                WHERE wh.project_id = ? ORDER BY wh.work_date
            """, (project_id,)).fetchall()
        else:
            rows = cursor.execute("""
                SELECT p.name as 项目, wh.work_date as 日期, w.name as 工人,
                       wh.hours as 工时, wh.work_type as 施工类型
                FROM work_hours wh JOIN workers w ON wh.worker_id = w.id
                JOIN projects p ON wh.project_id = p.id
                ORDER BY wh.work_date
            """).fetchall()
    elif table_name == 'curing':
        if project_id:
            rows = cursor.execute("""
                SELECT record_date as 日期, weather as 天气, temp_min as 最低温,
                       temp_max as 最高温, humidity as 湿度, curing_measure as 养护措施
                FROM curing_records WHERE project_id = ? ORDER BY record_date
            """, (project_id,)).fetchall()
        else:
            rows = cursor.execute("""
                SELECT p.name as 项目, record_date as 日期, weather as 天气,
                       temp_min as 最低温, temp_max as 最高温, curing_measure as 养护措施
                FROM curing_records JOIN projects p ON curing_records.project_id = p.id
                ORDER BY record_date
            """).fetchall()
    elif table_name == 'workers':
        rows = cursor.execute("""
            SELECT w.name as 姓名, w.role as 角色, w.phone as 电话,
                   w.hourly_rate as 时薪, t.name as 班组, w.notes as 备注
            FROM workers w LEFT JOIN teams t ON w.team_id = t.id
            ORDER BY w.name
        """).fetchall()
    elif table_name == 'suppliers':
        rows = cursor.execute("""
            SELECT s.name as 供应商名称, s.contact_person as 联系人, s.phone as 电话,
                   s.address as 地址, s.materials as 供应材料, s.rating as 评分, s.notes as 备注
            FROM suppliers s ORDER BY s.name
        """).fetchall()
    elif table_name == 'equipment':
        rows = cursor.execute("""
            SELECT e.name as 设备名称, e.type as 类型, e.model as 型号,
                   e.quantity as 数量, e.unit as 单位, e.status as 状态,
                   e.next_maintenance as 下次保养
            FROM equipment e ORDER BY e.name
        """).fetchall()
    else:
        conn.close()
        return ""

    conn.close()

    if not rows:
        return "无数据"

    # 转为CSV
    import io
    import csv
    output = io.StringIO()
    writer = csv.writer(output)
    # 头部
    writer.writerow([d for d in rows[0].keys()])
    # 数据
    for r in rows:
        writer.writerow([str(v) for v in r])

    return output.getvalue()


# ============================================================
# 数据导出
# ============================================================

def export_project_data(project_id: int) -> Dict:
    """导出项目的完整数据"""
    project = get_project(project_id)
    if not project:
        return {"error": "项目不存在"}

    conn = get_db()
    cursor = conn.cursor()

    # 日志
    logs = [dict(r) for r in cursor.execute(
        "SELECT * FROM daily_logs WHERE project_id = ? ORDER BY log_date",
        (project_id,)).fetchall()]

    # 检查清单
    checklist = [dict(r) for r in cursor.execute(
        "SELECT * FROM checklist_state WHERE project_id = ? ORDER BY id",
        (project_id,)).fetchall()]

    # 质量检测
    quality = [dict(r) for r in cursor.execute(
        "SELECT * FROM quality_tests WHERE project_id = ? ORDER BY id",
        (project_id,)).fetchall()]

    # 材料记录
    materials = [dict(r) for r in cursor.execute(
        "SELECT * FROM material_records WHERE project_id = ? ORDER BY record_date",
        (project_id,)).fetchall()]

    # 养护记录
    curing = [dict(r) for r in cursor.execute(
        "SELECT * FROM curing_records WHERE project_id = ? ORDER BY record_date",
        (project_id,)).fetchall()]

    # 预算
    budget = [dict(r) for r in cursor.execute(
        "SELECT * FROM budget_items WHERE project_id = ? ORDER BY id",
        (project_id,)).fetchall()]

    # 工时记录
    work_hours_data = [dict(r) for r in cursor.execute(
        "SELECT wh.*, w.name as worker_name FROM work_hours wh "
        "JOIN workers w ON wh.worker_id = w.id "
        "WHERE wh.project_id = ? ORDER BY wh.work_date",
        (project_id,)).fetchall()]

    conn.close()

    # 材料用量计算
    from materials_calc import calculate_all, calc_purchase_list
    calc = calculate_all(project['area'], project['base_thickness'], project['surface_thickness'])
    purchase = calc_purchase_list(calc['summary'])

    return {
        "export_version": "3.9.0",
        "exported_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "project": dict(project),
        "statistics": {
            "log_count": len(logs),
            "checklist_count": len(checklist),
            "checked_count": sum(1 for c in checklist if c['is_checked']),
            "quality_count": len(quality),
            "passed_count": sum(1 for q in quality if q['is_pass']),
            "material_records": len(materials),
            "curing_records": len(curing),
            "budget_items": len(budget),
            "work_hours": len(work_hours_data),
            "total_material_kg": calc['total_weight_kg'],
        },
        "daily_logs": logs,
        "checklist": checklist,
        "quality_tests": quality,
        "material_records": materials,
        "curing_records": curing,
        "budget_items": budget,
        "work_hours": work_hours_data,
        "material_calculation": calc,
        "purchase_list": purchase,
    }


def backup_database() -> str:
    """备份数据库，返回备份文件路径"""
    import shutil
    backup_dir = os.path.join(os.path.dirname(__file__), 'data', 'backups')
    os.makedirs(backup_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = os.path.join(backup_dir, f"yongyi_backup_{timestamp}.db")

    conn = get_db()
    conn.execute("VACUUM")
    conn.close()

    shutil.copy2(DB_PATH, backup_path)
    return backup_path


def restore_database(backup_path: str) -> Dict:
    """从备份文件恢复数据库

    参数:
        backup_path: 备份文件路径

    返回:
        {"status": "ok"} 或 {"error": "..."}
    """
    import shutil
    import os

    if not os.path.exists(backup_path):
        return {"error": f"备份文件不存在: {backup_path}"}

    if not backup_path.endswith('.db'):
        return {"error": "无效的备份文件"}

    if not os.path.isfile(backup_path):
        return {"error": "路径不是文件"}

    try:
        # 关闭当前连接
        conn = get_db()
        conn.close()

        # 先备份当前数据库（以防恢复出错）
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        pre_restore_backup = DB_PATH + f".pre_restore_{timestamp}"
        shutil.copy2(DB_PATH, pre_restore_backup)

        # 恢复
        shutil.copy2(backup_path, DB_PATH)

        # 验证恢复后的数据库
        test_conn = sqlite3.connect(DB_PATH)
        cursor = test_conn.cursor()
        cursor.execute("PRAGMA integrity_check")
        result = cursor.fetchone()[0]
        test_conn.close()

        if result != "ok":
            # 恢复失败，回滚
            shutil.copy2(pre_restore_backup, DB_PATH)
            return {"error": f"数据库完整性检查失败: {result}"}

        return {"status": "ok", "pre_restore_backup": pre_restore_backup}

    except Exception as e:
        return {"error": f"恢复失败: {str(e)}"}


def get_backup_list() -> List[Dict]:
    """获取备份文件列表"""
    backup_dir = os.path.join(os.path.dirname(__file__), 'data', 'backups')
    if not os.path.exists(backup_dir):
        return []

    backups = []
    for f in sorted(os.listdir(backup_dir), reverse=True):
        if f.endswith('.db') and f.startswith('yongyi_backup_'):
            path = os.path.join(backup_dir, f)
            size_kb = round(os.path.getsize(path) / 1024, 1)
            # 从文件名解析时间
            parts = f.replace('yongyi_backup_', '').replace('.db', '')
            backups.append({
                "filename": f,
                "path": path,
                "size_kb": size_kb,
                "created_at": parts
            })
    return backups


# ============================================================
# 施工照片管理
# ============================================================

import base64
import uuid

PHOTO_DIR = os.path.join(os.path.dirname(__file__), 'data', 'photos')


def _ensure_photo_dir():
    """确保照片存储目录存在"""
    os.makedirs(PHOTO_DIR, exist_ok=True)


def add_photo(project_id: int, phase: str, description: str,
              image_data: str) -> Dict:
    """
    添加施工照片

    参数:
        project_id: 项目ID
        phase: 施工阶段 (如: 基层处理, 抗裂砂浆施工, 面层施工, 打磨抛光, 密封固化)
        description: 照片描述
        image_data: base64编码的图片数据 (data:image/...;base64,...)

    返回:
        照片记录
    """
    _ensure_photo_dir()

    # 解析base64
    import re
    match = re.match(r'data:image/(\w+);base64,(.+)', image_data)
    if not match:
        # 如果没有data URL前缀，直接当纯base64处理
        fmt = 'png'
        b64_data = image_data
    else:
        fmt = match.group(1)
        b64_data = match.group(2)

    # 生成唯一文件名
    filename = f"{uuid.uuid4().hex}.{fmt}"
    filepath = os.path.join(PHOTO_DIR, filename)

    # 解码并保存
    try:
        img_bytes = base64.b64decode(b64_data)
        with open(filepath, 'wb') as f:
            f.write(img_bytes)
    except Exception as e:
        return {"error": f"图片解码失败: {str(e)}"}

    file_size = os.path.getsize(filepath)

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO photos (project_id, phase, description, filename, file_size)
        VALUES (?, ?, ?, ?, ?)
    """, (project_id, phase, description, filename, file_size))
    photo_id = cursor.lastrowid
    conn.commit()
    conn.close()

    return {
        "id": photo_id,
        "filename": filename,
        "file_size": file_size,
        "url": f"/api/photo/{photo_id}",
    }


def get_photos(project_id: int, phase: str = "") -> List[Dict]:
    """获取项目的照片列表"""
    conn = get_db()
    cursor = conn.cursor()
    if phase:
        rows = cursor.execute("""
            SELECT * FROM photos WHERE project_id = ? AND phase = ?
            ORDER BY created_at DESC
        """, (project_id, phase)).fetchall()
    else:
        rows = cursor.execute("""
            SELECT * FROM photos WHERE project_id = ?
            ORDER BY created_at DESC
        """, (project_id,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_photo_by_id(photo_id: int) -> Optional[Dict]:
    """获取单张照片信息"""
    conn = get_db()
    cursor = conn.cursor()
    row = cursor.execute("SELECT * FROM photos WHERE id = ?", (photo_id,)).fetchone()
    conn.close()
    return dict(row) if row else None


def delete_photo(photo_id: int) -> bool:
    """删除照片"""
    photo = get_photo_by_id(photo_id)
    if not photo:
        return False
    # 删除文件
    filepath = os.path.join(PHOTO_DIR, photo['filename'])
    if os.path.exists(filepath):
        os.remove(filepath)
    # 删除记录
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM photos WHERE id = ?", (photo_id,))
    conn.commit()
    conn.close()
    return True


# ============================================================
# 甘特图时间轴数据
# ============================================================

def get_timeline_data(project_id: int) -> Dict:
    """
    获取项目时间轴数据（用于甘特图）
    结合进度计划和实际日志生成时间轴
    """
    from materials_calc import generate_schedule

    project = get_project(project_id)
    if not project:
        return {"error": "项目不存在"}

    # 获取计划时间
    start_date = project.get('start_date', '')
    schedule = generate_schedule(start_date)

    # 获取实际施工日志
    logs = get_daily_logs(project_id)
    actual_dates = set()
    for log in logs:
        actual_dates.add(log['log_date'])

    # 获取检查清单进度
    checklist = get_checklist(project_id)
    phase_progress = {}
    for cat in checklist:
        phase = cat.get('phase', '')
        items = cat['items']
        checked = sum(1 for i in items if i['checked'])
        total = len(items)
        if phase:
            phase_progress[phase] = {
                "progress": round(checked / total * 100, 1) if total > 0 else 0,
                "checked": checked,
                "total": total,
            }

    # 构建时间轴
    timeline = []
    for phase in schedule:
        timeline.append({
            "phase": phase['phase'],
            "planned_start": phase['start'],
            "planned_end": phase['end'],
            "days": phase['days'],
            "actual_days": sum(1 for d in actual_dates if phase['start'] <= d <= phase['end']),
            "progress": phase_progress.get(phase['phase'], {}).get('progress', 0),
        })

    return {
        "project": project['name'],
        "start_date": start_date,
        "timeline": timeline,
        "total_planned_days": sum(p['days'] for p in timeline),
        "total_actual_days": len(actual_dates),
    }


# ============================================================
# 性能优化
# ============================================================

def optimize_database():
    """数据库性能优化: 创建索引 + VACUUM"""
    conn = get_db()
    cursor = conn.cursor()

    # 创建索引
    indexes = [
        "CREATE INDEX IF NOT EXISTS idx_checklist_project ON checklist_state(project_id)",
        "CREATE INDEX IF NOT EXISTS idx_checklist_phase ON checklist_state(phase)",
        "CREATE INDEX IF NOT EXISTS idx_logs_project ON daily_logs(project_id)",
        "CREATE INDEX IF NOT EXISTS idx_logs_date ON daily_logs(log_date)",
        "CREATE INDEX IF NOT EXISTS idx_quality_project ON quality_tests(project_id)",
        "CREATE INDEX IF NOT EXISTS idx_photos_project ON photos(project_id)",
        "CREATE INDEX IF NOT EXISTS idx_photos_phase ON photos(phase)",
        "CREATE INDEX IF NOT EXISTS idx_inventory_project ON material_records(project_id)",
        "CREATE INDEX IF NOT EXISTS idx_inventory_name ON material_records(material_name)",
        "CREATE INDEX IF NOT EXISTS idx_projects_status ON projects(status)",
        "CREATE INDEX IF NOT EXISTS idx_suppliers_name ON suppliers(name)",
        "CREATE INDEX IF NOT EXISTS idx_supplier_prices_supplier ON supplier_prices(supplier_id)",
        "CREATE INDEX IF NOT EXISTS idx_supplier_prices_material ON supplier_prices(material_name)",
        "CREATE INDEX IF NOT EXISTS idx_curing_project ON curing_records(project_id)",
        "CREATE INDEX IF NOT EXISTS idx_curing_date ON curing_records(record_date)",
        "CREATE INDEX IF NOT EXISTS idx_workers_team ON workers(team_id)",
        "CREATE INDEX IF NOT EXISTS idx_work_hours_project ON work_hours(project_id)",
        "CREATE INDEX IF NOT EXISTS idx_work_hours_worker ON work_hours(worker_id)",
        "CREATE INDEX IF NOT EXISTS idx_work_hours_date ON work_hours(work_date)",
        "CREATE INDEX IF NOT EXISTS idx_budget_project ON budget_items(project_id)",
        "CREATE INDEX IF NOT EXISTS idx_budget_category ON budget_items(category)",
        "CREATE INDEX IF NOT EXISTS idx_notifications_project ON notifications(project_id)",
        "CREATE INDEX IF NOT EXISTS idx_notifications_read ON notifications(is_read)",
        "CREATE INDEX IF NOT EXISTS idx_equipment_type ON equipment(type)",
        "CREATE INDEX IF NOT EXISTS idx_equipment_status ON equipment(status)",
        "CREATE INDEX IF NOT EXISTS idx_equip_usage_equip ON equipment_usage(equipment_id)",
        "CREATE INDEX IF NOT EXISTS idx_equip_usage_project ON equipment_usage(project_id)",
        "CREATE INDEX IF NOT EXISTS idx_acceptance_project ON acceptance_records(project_id)",
        "CREATE INDEX IF NOT EXISTS idx_acceptance_items_acc ON acceptance_items(acceptance_id)",
        "CREATE INDEX IF NOT EXISTS idx_safety_check_project ON safety_checks(project_id)",
        "CREATE INDEX IF NOT EXISTS idx_safety_check_date ON safety_checks(check_date)",
        "CREATE INDEX IF NOT EXISTS idx_safety_incident_project ON safety_incidents(project_id)",
        "CREATE INDEX IF NOT EXISTS idx_documents_project ON documents(project_id)",
        "CREATE INDEX IF NOT EXISTS idx_documents_type ON documents(doc_type)",
        "CREATE INDEX IF NOT EXISTS idx_mat_req_project ON material_requests(project_id)",
        "CREATE INDEX IF NOT EXISTS idx_mat_req_status ON material_requests(status)",
        "CREATE INDEX IF NOT EXISTS idx_req_items_request ON request_items(request_id)",
        "CREATE INDEX IF NOT EXISTS idx_subcontractors_status ON subcontractors(status)",
        "CREATE INDEX IF NOT EXISTS idx_templates_name ON project_templates(name)",
        "CREATE INDEX IF NOT EXISTS idx_consumption_project ON material_consumption(project_id)",
        "CREATE INDEX IF NOT EXISTS idx_consumption_material ON material_consumption(material_name)",
    ]
    for idx in indexes:
        cursor.execute(idx)

    # 分析优化
    cursor.execute("ANALYZE")

    conn.commit()
    conn.close()
    return {"indexes_created": len(indexes)}


def get_db_stats() -> Dict:
    """获取数据库统计信息"""
    conn = get_db()
    cursor = conn.cursor()

    stats = {}
    tables = ["projects", "daily_logs", "quality_tests", "checklist_state",
              "material_records", "photos", "suppliers", "supplier_prices",
              "curing_records", "workers", "teams", "work_hours",
              "budget_items", "notifications", "equipment", "equipment_usage",
              "acceptance_records", "acceptance_items",
              "safety_checks", "safety_incidents", "documents",
              "material_requests", "request_items", "subcontractors",
              "project_templates", "material_consumption"]
    for table in tables:
        count = cursor.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
        stats[table] = count

    # 数据库大小
    db_size = os.path.getsize(DB_PATH) if os.path.exists(DB_PATH) else 0
    stats["db_size_bytes"] = db_size
    stats["db_size_mb"] = round(db_size / 1024 / 1024, 2)

    conn.close()
    return stats


# ============================================================
# 数据导入
# ============================================================

def import_project(data: Dict) -> Dict:
    """
    从导出的数据导入项目（含所有关联数据）

    参数:
        data: export_project_data() 导出的字典

    返回:
        导入结果
    """
    project_data = data.get('project', {})
    if not project_data:
        return {"error": "无项目数据"}

    # 创建项目
    pid = create_project(
        name=project_data.get('name', '导入项目'),
        area=float(project_data.get('area', 100)),
        base_thickness=float(project_data.get('base_thickness', 50)),
        surface_thickness=float(project_data.get('surface_thickness', 15)),
        start_date=project_data.get('start_date', ''),
        location=project_data.get('location', ''),
    )

    conn = get_db()
    cursor = conn.cursor()

    import_count = {"logs": 0, "quality": 0, "materials": 0}

    # 导入施工日志
    for log in data.get('daily_logs', []):
        cursor.execute("""
            INSERT INTO daily_logs (project_id, log_date, weather, temperature,
                workers, work_content, materials_used, issues)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (pid, log.get('log_date', ''), log.get('weather', '晴'),
              log.get('temperature', ''), int(log.get('workers', 5)),
              log.get('work_content', ''), log.get('materials_used', ''),
              log.get('issues', '无')))
        import_count["logs"] += 1

    # 导入质量检测
    for qt in data.get('quality_tests', []):
        cursor.execute("""
            INSERT INTO quality_tests (project_id, test_id, test_name,
                standard_value, actual_value, is_pass)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (pid, qt.get('test_id', ''), qt.get('test_name', ''),
              qt.get('standard_value', ''), qt.get('actual_value', ''),
              1 if qt.get('is_pass') else 0))
        import_count["quality"] += 1

    # 导入材料记录
    for mat in data.get('material_records', []):
        cursor.execute("""
            INSERT INTO material_records (project_id, material_name,
                quantity_kg, quantity_packages, unit_price, total_cost, record_date)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (pid, mat.get('material_name', ''), float(mat.get('quantity_kg', 0)),
              float(mat.get('quantity_packages', 0)), float(mat.get('unit_price', 0)),
              float(mat.get('total_cost', 0)), mat.get('record_date', '')))
        import_count["materials"] += 1

    # 导入预算
    for bi in data.get('budget_items', []):
        cursor.execute("""
            INSERT INTO budget_items (project_id, category, planned_amount, actual_amount, description)
            VALUES (?, ?, ?, ?, ?)
        """, (pid, bi.get('category', ''), float(bi.get('planned_amount', 0)),
              float(bi.get('actual_amount', 0)), bi.get('description', '')))
        import_count["budget"] = import_count.get("budget", 0) + 1

    # 导入养护记录
    for cr in data.get('curing_records', []):
        cursor.execute("""
            INSERT INTO curing_records (project_id, record_date, weather,
                temp_min, temp_max, humidity, wind_level, curing_measure, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (pid, cr.get('record_date', ''), cr.get('weather', '晴'),
              float(cr.get('temp_min', 20)), float(cr.get('temp_max', 25)),
              float(cr.get('humidity', 60)), cr.get('wind_level', ''),
              cr.get('curing_measure', ''), cr.get('notes', '')))
        import_count["curing"] = import_count.get("curing", 0) + 1

    conn.commit()
    conn.close()

    return {
        "project_id": pid,
        "project_name": project_data.get('name', ''),
        "imported": import_count,
    }


# ============================================================
# 用户认证管理
# ============================================================

def add_user(username: str, password_hash: str, display_name: str,
             role: str = "worker", phone: str = "", email: str = "") -> int:
    """添加用户"""
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            INSERT INTO users (username, password_hash, display_name, role, phone, email)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (username, password_hash, display_name, role, phone, email))
        uid = cursor.lastrowid
        conn.commit()
        return uid
    except sqlite3.IntegrityError:
        return -1
    finally:
        conn.close()


def get_user_by_username(username: str) -> Optional[Dict]:
    """通过用户名获取用户"""
    conn = get_db()
    cursor = conn.cursor()
    user = cursor.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()
    conn.close()
    return dict(user) if user else None


def get_user_by_id(uid: int) -> Optional[Dict]:
    """通过ID获取用户"""
    conn = get_db()
    cursor = conn.cursor()
    user = cursor.execute("SELECT * FROM users WHERE id = ?", (uid,)).fetchone()
    conn.close()
    return dict(user) if user else None


def get_users() -> List[Dict]:
    """获取所有用户"""
    conn = get_db()
    cursor = conn.cursor()
    users = cursor.execute("""
        SELECT id, username, display_name, role, phone, email, is_active, last_login, created_at
        FROM users ORDER BY id
    """).fetchall()
    conn.close()
    return [dict(u) for u in users]


def update_user(uid: int, **kwargs) -> bool:
    """更新用户信息"""
    allowed = ['display_name', 'role', 'phone', 'email', 'is_active', 'password_hash']
    updates = {k: v for k, v in kwargs.items() if k in allowed}
    if not updates:
        return False
    set_clause = ", ".join(f"{k} = ?" for k in updates)
    values = list(updates.values()) + [uid]
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(f"UPDATE users SET {set_clause} WHERE id = ?", values)
    ok = cursor.rowcount > 0
    conn.commit()
    conn.close()
    return ok


def delete_user(uid: int) -> bool:
    """删除用户"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM users WHERE id = ?", (uid,))
    ok = cursor.rowcount > 0
    conn.commit()
    conn.close()
    return ok


def update_last_login(uid: int):
    """更新最后登录时间"""
    from datetime import datetime
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE users SET last_login = ? WHERE id = ?",
                   (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), uid))
    conn.commit()
    conn.close()


# ============================================================
# 环境监测记录
# ============================================================

def add_environment_record(project_id: int, record_date: str = "",
                           record_time: str = "", temperature: float = 0,
                           humidity: float = 0, base_moisture: float = None,
                           surface_temp: float = None, wind_speed: float = None,
                           weather_condition: str = "", recorder: str = "",
                           notes: str = "") -> int:
    """添加环境监测记录"""
    from datetime import date
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO environment_records
        (project_id, record_date, record_time, temperature, humidity,
         base_moisture, surface_temp, wind_speed, weather_condition, recorder, notes)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (project_id, record_date or str(date.today()), record_time,
          temperature, humidity, base_moisture, surface_temp, wind_speed,
          weather_condition, recorder, notes))
    rid = cursor.lastrowid
    conn.commit()
    conn.close()
    return rid


def get_environment_records(project_id: int, days: int = 30) -> List[Dict]:
    """获取环境监测记录"""
    conn = get_db()
    cursor = conn.cursor()
    records = cursor.execute("""
        SELECT * FROM environment_records
        WHERE project_id = ?
        ORDER BY record_date DESC, record_time DESC
        LIMIT ?
    """, (project_id, days * 10)).fetchall()
    conn.close()
    return [dict(r) for r in records]


def delete_environment_record(rid: int) -> bool:
    """删除环境记录"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM environment_records WHERE id = ?", (rid,))
    ok = cursor.rowcount > 0
    conn.commit()
    conn.close()
    return ok


def get_environment_stats(project_id: int) -> Dict:
    """获取环境统计数据"""
    conn = get_db()
    cursor = conn.cursor()
    stats = cursor.execute("""
        SELECT
            COUNT(*) as total_records,
            ROUND(AVG(temperature), 1) as avg_temp,
            ROUND(MIN(temperature), 1) as min_temp,
            ROUND(MAX(temperature), 1) as max_temp,
            ROUND(AVG(humidity), 1) as avg_humidity,
            ROUND(MIN(humidity), 1) as min_humidity,
            ROUND(MAX(humidity), 1) as max_humidity,
            ROUND(AVG(base_moisture), 1) as avg_base_moisture,
            ROUND(AVG(surface_temp), 1) as avg_surface_temp
        FROM environment_records
        WHERE project_id = ?
    """, (project_id,)).fetchone()
    conn.close()
    return dict(stats) if stats and stats['total_records'] else {
        "total_records": 0, "avg_temp": 0, "min_temp": 0, "max_temp": 0,
        "avg_humidity": 0, "min_humidity": 0, "max_humidity": 0,
        "avg_base_moisture": None, "avg_surface_temp": None
    }


# ============================================================
# 工序状态机
# ============================================================

# 标准工序顺序
PHASE_ORDER = [
    "基层处理", "抗裂砂浆施工", "基层养护", "面层界面处理",
    "面层浇筑", "面层养护", "打磨抛光", "密封固化", "最终验收"
]


def init_project_phases(project_id: int):
    """初始化项目工序（创建项目时调用）"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM project_phases WHERE project_id = ?", (project_id,))
    for i, name in enumerate(PHASE_ORDER, 1):
        cursor.execute("""
            INSERT INTO project_phases (project_id, phase_name, phase_order, status)
            VALUES (?, ?, ?, ?)
        """, (project_id, name, i, 'pending'))
    conn.commit()
    conn.close()


def get_project_phases(project_id: int) -> List[Dict]:
    """获取项目工序状态"""
    conn = get_db()
    cursor = conn.cursor()
    phases = cursor.execute("""
        SELECT * FROM project_phases
        WHERE project_id = ?
        ORDER BY phase_order
    """, (project_id,)).fetchall()
    conn.close()
    if not phases:
        return [{"phase_name": p, "status": "pending", "phase_order": i+1}
                for i, p in enumerate(PHASE_ORDER)]
    return [dict(p) for p in phases]


VALID_PHASE_TRANSITIONS = {
    "pending": "in_progress",
    "in_progress": "completed",
    "completed": "in_progress",  # 可以回退修改
}


def update_phase_status(project_id: int, phase_name: str, new_status: str) -> Dict:
    """更新工序状态（带流转校验）

    状态转换规则:
    - 前置工序必须completed才能开始当前工序
    - 不能跳过工序
    """
    if phase_name not in PHASE_ORDER:
        return {"ok": False, "error": f"未知工序: {phase_name}"}

    if new_status not in VALID_PHASE_TRANSITIONS:
        return {"ok": False, "error": f"无效状态: {new_status}"}

    conn = get_db()
    cursor = conn.cursor()

    # 获取当前工序
    phase = cursor.execute("""
        SELECT * FROM project_phases
        WHERE project_id = ? AND phase_name = ?
    """, (project_id, phase_name)).fetchone()

    if not phase:
        # 如果数据库没有，初始化
        init_project_phases(project_id)
        phase = cursor.execute("""
            SELECT * FROM project_phases
            WHERE project_id = ? AND phase_name = ?
        """, (project_id, phase_name)).fetchone()
        if not phase:
            conn.close()
            return {"ok": False, "error": "工序初始化失败"}

    phase = dict(phase)
    current_status = phase['status']
    phase_order = phase['phase_order']

    # 校验状态转换合法性
    allowed = VALID_PHASE_TRANSITIONS.get(current_status, [])
    if new_status not in allowed and new_status != current_status:
        conn.close()
        return {"ok": False, "error": f"工序状态不能从'{current_status}'转换为'{new_status}'"}

    # 前置工序必须已完成（开始新工序时校验）
    if new_status == "in_progress" and current_status == "pending":
        prev_phase = cursor.execute("""
            SELECT status FROM project_phases
            WHERE project_id = ? AND phase_order = ?
        """, (project_id, phase_order - 1)).fetchone()
        if prev_phase and prev_phase['status'] != 'completed':
            conn.close()
            return {"ok": False, "error": f"前置工序未完成，不能开始'{phase_name}'"}

    from datetime import datetime
    now = datetime.now().strftime('%Y-%m-%d %H:%M')

    if new_status == "in_progress":
        cursor.execute("""
            UPDATE project_phases SET status = ?, started_at = ?
            WHERE project_id = ? AND phase_name = ?
        """, (new_status, now, project_id, phase_name))
    elif new_status == "completed":
        cursor.execute("""
            UPDATE project_phases SET status = ?, completed_at = ?
            WHERE project_id = ? AND phase_name = ?
        """, (new_status, now, project_id, phase_name))
    else:
        cursor.execute("""
            UPDATE project_phases SET status = ?
            WHERE project_id = ? AND phase_name = ?
        """, (new_status, project_id, phase_name))

    conn.commit()

    # 更新项目整体进度
    all_phases = cursor.execute("""
        SELECT status FROM project_phases
        WHERE project_id = ? ORDER BY phase_order
    """, (project_id,)).fetchall()

    completed_count = sum(1 for p in all_phases if p['status'] == 'completed')
    total_count = len(all_phases)

    from datetime import datetime
    if completed_count == total_count:
        cursor.execute("UPDATE projects SET status = '已完成', updated_at = ? WHERE id = ?",
                       (datetime.now().strftime('%Y-%m-%d %H:%M'), project_id))
    elif completed_count > 0:
        cursor.execute("UPDATE projects SET status = '进行中', updated_at = ? WHERE id = ?",
                       (datetime.now().strftime('%Y-%m-%d %H:%M'), project_id))

    conn.commit()
    conn.close()

    return {
        "ok": True,
        "phase_name": phase_name,
        "new_status": new_status,
        "progress": f"{completed_count}/{total_count}",
    }


def check_database_integrity() -> Dict:
    """检查数据库完整性"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("PRAGMA integrity_check")
        result = cursor.fetchone()[0]
        conn.close()
        return {"status": "ok" if result == "ok" else "error", "detail": result}
    except Exception as e:
        return {"status": "error", "detail": str(e)}


def get_database_stats_detailed() -> Dict:
    """获取详细的数据库统计"""
    conn = get_db()
    cursor = conn.cursor()
    stats = {
        "db_size_mb": round(os.path.getsize(DB_PATH) / (1024 * 1024), 2) if os.path.exists(DB_PATH) else 0,
        "table_count": 0,
        "index_count": 0,
    }
    try:
        cursor.execute("SELECT COUNT(*) FROM sqlite_master WHERE type='table'")
        stats["table_count"] = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM sqlite_master WHERE type='index'")
        stats["index_count"] = cursor.fetchone()[0]
    except:
        pass
    conn.close()
    return stats


# ============================================================
# Excel 导出 (v4.6.0)
# ============================================================

def export_to_excel(project_id: int = 0, table_name: str = "projects") -> Optional[bytes]:
    """导出项目数据为Excel格式

    参数:
        project_id: 项目ID (0=所有项目)
        table_name: 表名 (projects/logs/quality/materials/workers/suppliers/equipment)

    返回:
        Excel文件字节流
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    conn = get_db()
    cursor = conn.cursor()
    from io import BytesIO
    
    filename_map = {
        'projects': '项目汇总', 'logs': '施工日志', 'quality': '质量检测',
        'materials': '材料记录', 'workers': '工人花名册',
        'suppliers': '供应商清单', 'equipment': '设备清单',
    }
    
    wb = Workbook()
    ws = wb.active
    ws.title = filename_map.get(table_name, table_name)

    # 样式定义
    header_font = Font(name='Microsoft YaHei', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1A5276', end_color='1A5276', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    cell_alignment = Alignment(vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    def _write_header(headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    def _write_rows(data_rows):
        for r, row in enumerate(data_rows, 2):
            for c, val in enumerate(row, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.alignment = cell_alignment
                cell.border = thin_border

    def _auto_width():
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    # 中文字符按2倍宽度计算
                    val = str(cell.value)
                    length = sum(2 if ord(c) > 127 else 1 for c in val)
                    max_len = max(max_len, length)
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    try:
        if table_name == 'projects':
            if project_id:
                rows = cursor.execute("""
                    SELECT id AS '项目ID', name AS '项目名称', area AS '面积(m²)',
                           base_thickness AS '基层厚度(mm)', surface_thickness AS '面层厚度(mm)',
                           status AS '状态', created_at AS '创建时间'
                    FROM projects WHERE id = ? ORDER BY id
                """, (project_id,)).fetchall()
            else:
                rows = cursor.execute("""
                    SELECT id AS '项目ID', name AS '项目名称', area AS '面积(m²)',
                           base_thickness AS '基层厚度(mm)', surface_thickness AS '面层厚度(mm)',
                           status AS '状态', created_at AS '创建时间'
                    FROM projects ORDER BY id
                """).fetchall()
            if rows:
                _write_header(list(rows[0].keys()))
                _write_rows([list(r) for r in rows])

        elif table_name == 'logs':
            if project_id:
                rows = cursor.execute("""
                    SELECT log_date AS '日期', weather AS '天气', temperature AS '温度(℃)',
                           workers AS '工人数', work_content AS '施工内容',
                           materials_used AS '材料使用', issues AS '问题'
                    FROM daily_logs WHERE project_id = ? ORDER BY log_date
                """, (project_id,)).fetchall()
            else:
                rows = cursor.execute("""
                    SELECT p.name AS '项目', l.log_date AS '日期', l.weather AS '天气',
                           l.temperature AS '温度(℃)', l.workers AS '工人数',
                           l.work_content AS '施工内容', l.issues AS '问题'
                    FROM daily_logs l JOIN projects p ON l.project_id = p.id ORDER BY l.log_date
                """).fetchall()
            if rows:
                _write_header(list(rows[0].keys()))
                _write_rows([list(r) for r in rows])

        elif table_name == 'quality':
            if project_id:
                rows = cursor.execute("""
                    SELECT test_name AS '检测项目', standard_value AS '标准值',
                           actual_value AS '实测值',
                           CASE WHEN is_pass = 1 THEN '合格' ELSE '不合格' END AS '结果',
                           test_date AS '检测日期'
                    FROM quality_tests WHERE project_id = ? ORDER BY test_date
                """, (project_id,)).fetchall()
            else:
                rows = cursor.execute("""
                    SELECT p.name AS '项目', q.test_name AS '检测项目',
                           q.standard_value AS '标准值', q.actual_value AS '实测值',
                           CASE WHEN q.is_pass = 1 THEN '合格' ELSE '不合格' END AS '结果',
                           q.test_date AS '检测日期'
                    FROM quality_tests q JOIN projects p ON q.project_id = p.id ORDER BY q.test_date
                """).fetchall()
            if rows:
                _write_header(list(rows[0].keys()))
                _write_rows([list(r) for r in rows])

        elif table_name == 'materials':
            if project_id:
                rows = cursor.execute("""
                    SELECT material_name AS '材料名称', quantity_kg AS '数量(kg)',
                           quantity_packages AS '包装数', unit_price AS '单价(元)',
                           total_cost AS '总价(元)', record_date AS '记录日期'
                    FROM material_records WHERE project_id = ? ORDER BY record_date
                """, (project_id,)).fetchall()
            else:
                rows = cursor.execute("""
                    SELECT p.name AS '项目', m.material_name AS '材料名称',
                           m.quantity_kg AS '数量(kg)', m.unit_price AS '单价(元)',
                           m.total_cost AS '总价(元)', m.record_date AS '记录日期'
                    FROM material_records m JOIN projects p ON m.project_id = p.id ORDER BY m.record_date
                """).fetchall()
            if rows:
                _write_header(list(rows[0].keys()))
                _write_rows([list(r) for r in rows])

        elif table_name == 'workers':
            if project_id:
                rows = cursor.execute("""
                    SELECT name AS '姓名', phone AS '电话', role AS '工种',
                           hourly_rate AS '时薪(元)', notes AS '备注'
                    FROM workers WHERE team_id IN (
                        SELECT id FROM teams WHERE project_id = ?
                    ) ORDER BY name
                """, (project_id,)).fetchall()
            else:
                rows = cursor.execute("""
                    SELECT name AS '姓名', phone AS '电话', role AS '工种',
                           hourly_rate AS '时薪(元)', notes AS '备注'
                    FROM workers ORDER BY name
                """).fetchall()
            if rows:
                _write_header(list(rows[0].keys()))
                _write_rows([list(r) for r in rows])

        elif table_name == 'suppliers':
            rows = cursor.execute("""
                SELECT name AS '供应商名称', contact_person AS '联系人',
                       phone AS '电话', materials AS '供应材料',
                       rating AS '评分', notes AS '备注'
                FROM suppliers ORDER BY name
            """).fetchall()
            if rows:
                _write_header(list(rows[0].keys()))
                _write_rows([list(r) for r in rows])

        elif table_name == 'equipment':
            rows = cursor.execute("""
                SELECT name AS '设备名称', type AS '类型', model AS '型号',
                       quantity AS '数量', status AS '状态',
                       purchase_date AS '购置日期'
                FROM equipment ORDER BY name
            """).fetchall()
            if rows:
                _write_header(list(rows[0].keys()))
                _write_rows([list(r) for r in rows])

        else:
            conn.close()
            return None

        _auto_width()
        conn.close()

        # 冻结首行
        ws.freeze_panes = 'A2'

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    except Exception as e:
        conn.close()
        raise e
